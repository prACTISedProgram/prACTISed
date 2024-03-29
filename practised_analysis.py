#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# practised_analysis.py calculates the Kd value from ACTIS experimental
# data in an Excel working file formatted for practised.py. Generates
# separagram and binding isotherm graphs

# Copyright (C) 2022  Jessica Latimer

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.


def dataanalysis(fileName):

        ## Part 1 - Importing the required libraries and sub-libraries required below
        import argparse                                   
        import pathlib
        import sys
        import pandas as pd
        import matplotlib.pyplot as plt
        import numpy as np
        from openpyxl import load_workbook
        from openpyxl import Workbook
        from openpyxl.utils import FORMULAE
        from openpyxl.utils import get_column_letter, column_index_from_string
        import math
        from scipy.optimize import curve_fit
        from datetime import date
        import os


        ## Part 3 - Locating the raw data file and establishing important inputs                         
        if not pathlib.Path(fileName).is_file():
                print("Given file '%s' is not a file or does not exist." % fileName)
                exit(-1)

        name = pathlib.PurePath(fileName).name

        # Set explicitly the engine to use openpyxl, otherwise it might use xlrd, which has removed support for Excel's xlsx format (only supports the old binary format)
        data = pd.read_excel(fileName, engine='openpyxl')
        inputBook = load_workbook(fileName, data_only=True)         
        inputBooknames = inputBook.sheetnames


        # Confirm Inputs sheet with correct formatting before preceding
        if "Inputs" in inputBook.sheetnames:                                            
            idealSheet = inputBook["Inputs"]
            percentage = float(idealSheet.cell(15,2).value)/100
            numberOfConcs = int(idealSheet.cell(10,2).value)
            injectionTime = float(idealSheet.cell(3,2).value)
            ligandConc = float(idealSheet.cell(11,2).value)
            
            proteinName = str(idealSheet.cell(8,2).value)
            dataType = str(idealSheet.cell(12,2).value)
            peakDet = str(idealSheet.cell(16,2).value)

            
            subdirect = "%s_graphs" % os.path.splitext(fileName)[0]
            if os.path.exists(subdirect)== False:
                    os.mkdir(subdirect)
                    
            elif os.path.exists(subdirect)== True:
                    for duplicate in range(2,9):
                            temp = "%s_%d" % (subdirect, duplicate)
                            
                            if os.path.exists(temp)== False:
                                    subdirect=temp
                                    os.mkdir(subdirect)
                                    break
            
        elif "Inputs" not in inputBook.sheetnames:  
                sys.exit("Input file Formatting Error: The script expects inputdata.xlsx to be in a certain format, see provided idealinputs.xlsx as an example.") 

        # Temporary variables
        concentration = []
        signal = []
        stddev = []
        relstddev = []
        Rvalue = []
        Rstddev = []
        relRstddev = []
        graphs = []
        graphNames = []
        forDF = [concentration,signal,stddev,relstddev,Rvalue,Rstddev,relRstddev]
        DFnames = ["Conc","Avg Sig (S)","S Std Dev","S Rel Std Dev","R value","R Std Dev", "R Rel Std Dev"]

        # If programmatic determination of peak use first run at specified concentration to calculate peak time and time window
        if peakDet == "P":
                windowCalcConc = float(idealSheet.cell(18,2).value)
                windowCalcConcS = "%s %s" % (windowCalcConc, idealSheet.cell(1,5).value.partition(" ")[2])

                data = pd.read_excel(fileName, sheet_name=windowCalcConcS, engine='openpyxl')
                data = data.dropna(how='all')
                xvalues = data['raw time']
                yvalues = data.iloc[:,1]

                # Convert raw time to propagation time
                yvalues = yvalues[xvalues >= injectionTime]
                xvalues = xvalues[xvalues >= injectionTime]
                xvalues = xvalues.subtract(injectionTime)

                peakSignal = max(yvalues)
                peakIndex = yvalues[yvalues == peakSignal].index
                peakTime = float(xvalues[peakIndex])


        # Determine absolute max signal value and max number of runs (used to set graph parameters)
        maxSig = 0
        maxRuns = 0
        manualPeaks = []
        for x in range(1,int(numberOfConcs)+1):         
                conc1 = idealSheet.cell(x,5).value
                data = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
                data = data.dropna(how='all')
                data2 = data.iloc[:,1:]
                currentMaxRuns = len(data2.columns)
                colMax = data.max()
                currentMax = colMax.max()

                if currentMax > maxSig:
                        maxSig = currentMax
                if currentMaxRuns > maxRuns:
                    maxRuns = currentMaxRuns

                 # If manual determination, extract the manually set peak time for given concentrations
                if peakDet == "M":
                        xvalues = data.iloc[:,0]
                        manualTimes = idealSheet.cell(17,2).value.split(",")
                        peakIndex = xvalues.searchsorted(float(manualTimes[x-1]), side='left')
                        peakTime = xvalues[peakIndex-1]
                        peakTime = peakTime
                        manualPeaks.append(peakTime)

                        windowCalcConc = 1^-25
        
        ## Part 4 - Calculating signal information for each concentration and generating separagram graphs
        for x in range(1,numberOfConcs+1):         
            conc1 = idealSheet.cell(x,5).value
            
            avgSigConc = 0
            avgSigsRun = []


            # Reading in the whole data frame (= all runs for one particular concentration) and dropping all lines that are blank, i.e. that would produce "NaN"s
            data = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
            data = data.dropna(how='all')
            numberOfRuns = len(data.columns)


            # Calculate background signals for each run
            for col in range(1, numberOfRuns):

                runName = data.columns[col]

                xvalues = data['raw time']
                yvalues = data[runName]
                minTime = xvalues[0]
                
                # All y-values before the injection time are considered background signal 
                background_yvalues = yvalues[xvalues < injectionTime]
                background_average = np.average(background_yvalues)
                background_stdev = np.std(background_yvalues)

                # Convert raw time to propagation time
                yvalues = yvalues[xvalues >= injectionTime]
                xvalues = xvalues[xvalues >= injectionTime]
                xvalues = xvalues.subtract(injectionTime)

                if peakDet == "M":
                        peakTime = manualPeaks[x-1]

                # Set time window parameters using peak time and determine the average signal within window for each run             
                windowLow = float(peakTime - (percentage * peakTime))
                windowHigh = float(peakTime + (percentage * peakTime))

                if percentage > 0:
                        windowIndex = xvalues.between(windowLow,windowHigh, inclusive='both')
                        windowTimes = xvalues[windowIndex]
                        windowSignals = yvalues[windowIndex]
                        
                elif percentage == 0:
                        closestIndex = xvalues.searchsorted(float(peakTime), side='left')
                        if xvalues[closestIndex] - peakTime >= 0.5:
                                sg.popup_ok('prACTISed cancelled \n \nError: No time %s found in %s run %s. \nPlease try again with same times for all experimental runs.' % (peakTime, conc1, col))
                                return False
                
                        windowSignals = yvalues[closestIndex]

                avgSigsRun.append(np.mean(windowSignals))


                # Graph the signal for each run and with time window indicated
                plt.plot(xvalues, yvalues)
                

            # Appending a figure with all experimental runs for concentration
            plt.xlabel('Propagation time (s)', fontweight='bold',fontsize=13)
            if dataType == "MS":
                    plt.ylabel('MS intensity (a.u.)', fontweight='bold', fontsize=13)
            elif dataType == "F":
                    plt.ylabel('Fluorescence (a.u.)', fontweight='bold', fontsize=13)
            plt.text(minTime,maxSig*1.05, r"[%s]$\mathbf{_0}$ = %s" % (proteinName, conc1), fontweight='bold', fontsize=13)
            plt.vlines(windowLow, 0, maxSig*1.05, linestyles='dashed',color='gray')
            plt.vlines(windowHigh, 0, maxSig*1.05, linestyles='dashed',color='gray')
            plt.savefig("%s/%s.png" % (subdirect, conc1))
            graphs.append(plt.figure())
            plt.clf()

            
            # Calculating average signal for each concentration, stdev and relative stdev
            avgSigConc = np.average(avgSigsRun)
            avgSigConc_stdev = np.std(avgSigsRun)
            avgSigConc_relstdev = (avgSigConc_stdev/avgSigConc)*100

            concentration.append(conc1)
            signal.append(avgSigConc)
            stddev.append(avgSigConc_stdev)
            relstddev.append(avgSigConc_relstdev)

            if peakDet == "P":
                if float(conc1.partition(" ")[0]) < windowCalcConc:
                        Rvalue.append(None)
                        Rstddev.append(None)
                        relRstddev.append(None)

            
        # Generate legend for run colors
        if maxRuns > 1:
                legendFig = plt.figure("Legend plot")
                for line in range(1, maxRuns):
                        plt.plot(line, np.sin(line), label='Run %s' % line)
                plt.axis('off')
                plt.title('Reference Legend')
                legendFig.legend(loc='center')
                legendFig.savefig('%s/legend.png' % (subdirect))
                plt.clf()
       
        # Graphing separagrams for the first run for every concentration
        lastConc = 0
        for x in range(1,int(numberOfConcs)+1):         
            conc1 = idealSheet.cell(x,5).value

            # Reading in the whole data frame (= all runs for one particular concentration) and dropping all lines that are blank, i.e. that would produce "NaN"s
            data = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
            data = data.dropna(how='all')
            xvalues = data['raw time']
            yvalues = data.iloc[:, 1]

            # Convert raw time to propagation time
            yvalues = yvalues[xvalues >= injectionTime]
            xvalues = xvalues[xvalues >= injectionTime]
            xvalues = xvalues.subtract(injectionTime)

            if peakDet == "M":
                peakTime = manualPeaks[x-1]

            checkConc = float(conc1.partition(" ")[0])
                
            if checkConc == 0 or checkConc == windowCalcConc or x == numberOfConcs or any(abs(yvalues[xvalues==peakTime]-lastConc)>=0.2*maxSig):
                p = plt.plot(xvalues, yvalues, label= '%s' % conc1)
                lastConc = yvalues[xvalues==peakTime]

        plt.xlabel('Propagation time (s)', fontweight='bold', fontsize=13)
        if dataType == "MS":
                plt.ylabel('MS intensity (a.u.)', fontweight='bold', fontsize=13)
        elif dataType == "F":
                plt.ylabel('Fluorescence (a.u.)', fontweight='bold', fontsize=13)
        plt.text(minTime, maxSig*1.05, r"[%s]$\mathbf{_0}$"  % (proteinName), fontweight='bold', fontsize=13)
        plt.legend(fontsize=13)
        if peakDet == "P":
                plt.vlines(windowLow, 0, maxSig*1.05, linestyles='dashed',color='gray')
                plt.vlines(windowHigh, 0, maxSig*1.05, linestyles='dashed',color='gray')
        elif peakDet == "M":
                plt.vlines(0, maxSig, maxSig*1.05, linestyles='dashed',color='white')

        # Appending a figure with first experimental run for all concentrations
        plt.savefig("%s/allconcentration.png" % subdirect)           # save separagram graph
        graphs.append(plt.figure())
        plt.clf()


        ## Part 5 - Calculate R values and standard deviation of R values for each concentration
        
        LowProt_sig = signal[len(Rvalue)]
        HighProt_sig = signal[len(concentration)-1]
        LowProt_stddev = stddev[len(Rvalue)]
        HighProt_stdDev = stddev[len(concentration)-1]


        for y in range(0, len(concentration)):
                conc2 = idealSheet.cell(y+1,5).value
                if float(conc2.partition(" ")[0]) >= windowCalcConc:
                        avgSigConc_R = (signal[y] - HighProt_sig)/ (LowProt_sig - HighProt_sig) 
                        Rvalue.append(avgSigConc_R)

                        avgSiglConc_Rstddev = ( 1/(LowProt_sig - HighProt_sig) * math.sqrt( (stddev[y]**2) + ((signal[y] - LowProt_sig)/ (LowProt_sig - HighProt_sig) * HighProt_stdDev)**2 +
                                                                                       ((HighProt_sig - signal[y])/(LowProt_sig - HighProt_sig) * LowProt_stddev)**2))
                        Rstddev.append(avgSiglConc_Rstddev)
                        avgSiglConc_relRstddev = (avgSiglConc_Rstddev/avgSigConc_R)*100
                        relRstddev.append(avgSiglConc_relRstddev)
  
        
        ## Part 6 - Plotting the binding isotherm R vs P[0] with curve of best fit
        # Convert concentration strings to floats
        concs = []
        for element in range(0, len(concentration)):
                num = concentration[element].partition(" ")[0]
                if float(num) >= windowCalcConc:
                        concs.append(float(num))
        unit = concentration[0].partition(" ")[2]
        

        Rvalue_drop = [i for i in Rvalue if i is not None]
        Rstddev_drop = [i for i in Rstddev if i is not None]
        
        # Plotting data points for each concentration
        plt.scatter(concs, Rvalue_drop, c='white', edgecolor='black', label="R", zorder=10)
        plt.errorbar(concs, Rvalue_drop, yerr = Rstddev_drop, linestyle="none", ecolor = 'black', elinewidth=1, capsize=2, capthick=1, zorder=0)
        plt.xscale("log")
                      
        # Define the Levenberg Marquardt algorithm
        def LevenMarqu(x,a):          
            return -((a + x - ligandConc)/(2*ligandConc)) + ((((a + x - ligandConc)/(2*ligandConc))**2) + (a/ligandConc))**(0.5)

        # Curve fitting and plotting curve of best fit    
        popt, pcov = curve_fit(LevenMarqu, concs, Rvalue_drop)
        error = np.sqrt(np.diag(pcov))

        step=0
        if concs[step] == float(0):
                step = 1

        xFit = np.arange(0.0, max(concs), concs[step])
        plt.plot(xFit, LevenMarqu(xFit, popt), linewidth=1.5, color='black', label="Best Fit")
        plt.text((concs[step]), 0.2, r'K$\mathbf{_d}$ = %.3g ± %.3g %s' % (popt, error, unit), fontweight='bold', fontsize=13)
        plt.ylabel('R', fontweight='bold', fontsize=13)
        plt.xlabel(r'[%s]$\mathbf{_0}$ (%s)' % (proteinName,unit), fontweight='bold',fontsize=13)
        plt.xscale("log")
        plt.legend(fontsize=13)
        plt.savefig("%s/bindingisotherm.png" % subdirect)           # save binding isotherm graph
        graphs.append(plt.figure())
        plt.close()

        # Statistics
        residuals = Rvalue_drop - LevenMarqu(concs, *popt)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((Rvalue_drop - np.mean(Rvalue_drop))**2)
        r_squared = 1 - (ss_res/ss_tot)

        chiSquared = sum((((Rvalue_drop - LevenMarqu(concs, *popt))**2) / LevenMarqu(concs, *popt)))


        ## Part 7 - Returning summary data and graphs
        # Summary dataframe of average signal per concentration with standard deviation, relative standard deviation, R value and standard deviation
        df = pd.DataFrame (forDF).transpose()
        df.columns = DFnames

        # Create new output sheet in input Excel file with summary data
        writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
        writer.book = inputBook
        df.to_excel(writer, sheet_name = "Outputs", float_format='%e', index=False, startcol=3, engine = 'openpyxl')        # does not overwrite if a sheet named Outputs already exists
        writer.save()
        writer.close()

        outputSheet = inputBook.worksheets[len(inputBook.sheetnames)-1]

        # Duplicate input sheet information onto output sheet for reproducibility
        for r in range(1, 19):
                for c in range (1, 3):
                        outputSheet.cell(row=r, column=c).value = idealSheet.cell(row=r, column=c).value

        
        # Include Kd, R² and χ² values on output sheet
        outputSheet["L1"] = "Kd: %.4f ± %.4f %s" % (popt,error,unit)
        outputSheet["L2"] = "R²: %.4f" % (r_squared)
        outputSheet["L3"] = "χ²: %.4f" % (chiSquared)

        inputBook.save(fileName)
        
        plt.close('all')

        return subdirect


