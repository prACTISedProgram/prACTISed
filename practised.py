#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# practised.py serves as the GUI for extracting ACTIS titration data
# users specify experimental parameters and practised.py calls
# assosciated modules to calculate Kd and return separagrams and
# binding isotherm graph. See READ.ME for information on Excel working 
# file preparation

# Copyright (C) 2022  Jessica Latimer

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.



import PySimpleGUI as sg

import os
import pathlib
import glob
import sys

from PIL import Image, ImageTk

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from scipy.optimize import curve_fit
from scipy import interpolate
from datetime import date
import time
from natsort import natsorted
from fpdf import FPDF
import webbrowser
from practised_working import workingfileprep
from practised_analysis import dataanalysis
from practised_compensation import compensate
from practised_pdfReport import report
from practised_validate import *
from practised_knuteon import *
import practised_pwexplode 

sg.theme('DarkBlue3')

# Building input column for GUI    
inputCol = [
    [sg.Frame('Fluidics Experimental Parameters', 
              [[sg.Text('Propagation flow rate', size=(27,1), key ='propFlow'),
                sg.Input('µL/min', size=(14,1), key='propFlow_val')],
               [sg.Text('Injection flow rate', size=(27,1),key='injectFlow'),
                sg.Input('µL/min', size=(14,1), key='injectFlow_val')],
               [sg.Text('Injection time (s) *', size=(27,1), key='injectTime'),
                sg.Input(size=(14,1), key='injectTime_val')],
               [sg.Text('Separation capillary length', size=(27,1), key='sepLength'),
                sg.Input('cm', size=(14,1), key='sepLength_val')],
               [sg.Text('Separation capillary diameter', size=(27,1), key='sepDiam'),
                sg.Input('µm', size=(14,1), key='sepDiam_val')],
               [sg.Text('Injection loop length', size=(27,1), key='injectLength'),
                sg.Input('cm', size=(14,1), key='injectLength_val')],
               [sg.Text('Injection loop diameter', size=(27,1), key='injectDiam'),
                sg.Input('µm', size=(14,1), key='injectDiam_val')]],
              key = 'fluidParam')],


    [sg.Frame('Concentrations', 
              [[sg.Text('Protein name *', size=(27,1), key ='protName'),
                sg.Input(size=(14,1), key='protName_val')],
               [sg.Text('Ligand name', size=(27,1), key='ligName'),
                sg.Input(size=(14,1), key='ligName_val')],
               [sg.Text(u'Initial ligand concentration [L]\u2080 *    \n (same units as [P]\u2080)', key='ligConc'),
                sg.Input(size=(14,1), key='ligConc_val')]],
              key='concFrame')],
    
    [sg.Frame('Data Analyis Parameters',
              [[sg.Text('Type of data *', size=(27,1), key ='dataType'),
                sg.Radio('Fluoresence', 'dataChoice', key ='dataF', enable_events=True, default=True)],
               [sg.Text('', size=(27,1), key ='blank'),
                sg.Radio('Mass Spec', 'dataChoice', key ='dataMS', enable_events=True)],
               [sg.Text('Compensation procedure *           \n (recommended for MS data)', key ='comp'),
                sg.Radio('Yes', 'compChoice', key ='compYes', enable_events=True)],
               [sg.Text('', size=(27,1), key ='blank'),
                sg.Radio('No', 'compChoice', key ='compNo', enable_events=True, default=True)],
               [sg.Text(u'[P]\u2080 reference for normalization *', size=(27,1), key='normConc', visible=False),
                sg.Input(size=(14,1), key='normConc_val', visible=False)],
               [sg.Text('Window width (%) *', size=(27,1), key='window'),
                sg.Input('2', size=(14,1), key='window_val')],
               [sg.Text('Determination of peak *', size=(27,1), key ='peak'),
                sg.Radio('Manual', 'peakChoice', key ='peakM', enable_events=True)],
               [sg.Text('', size=(27,1), key ='blank'),
                sg.Radio('Programmatic', 'peakChoice', key ='peakP', enable_events=True, default=True)],
               [sg.Text(u'Peaks for concentrations [P]\u2080 * \n (ascending [P]\u2080, separated by commas)', key='manPeaks', visible=False)],
               [sg.Input(size=(43,1), key='manPeaks_val', visible=False)],
               [sg.Text(u'Specify [P]\u2080 used to determine peak *', size=(30,1), key='progPeak', visible=True),
                sg.Input(size=(11,1), key='progPeak_val', visible=True)]],
              key='analysisParam')],
    
    [sg.Text('* Required field', key = 'req'),]
    ]


# Builing output column for GUI        
outputCol= [
    [sg.Table(values=[],headings=['prACTISed output'], key='Kd', hide_vertical_scroll=True, def_col_width=20, auto_size_columns=False)],

    [sg.Table(values=[], headings=['Conc','Avg Sig (S)', 'S Std Dev', 'S Rel Std Dev', 'R value', 'R Std Dev', 'R Rel Std Dev'], key='summary', def_col_width=10,auto_size_columns=False)],

    [sg.Frame('Graphs',
              [[sg.Image(key='graphImage')],
               [sg.Button('Prev', key='back'),
                sg.Button('Next', key='fwd')]]
               )]           
    ]

# Building file path column for GUI
filePathCol= [
    [sg.Text('File path *', key ='filePath'),
     sg.Input(size=(50,1), key='filePath_val'),
     sg.Button('Validate', key='validate')],
    [sg.Text('Working File Entered', key='work', visible=False),
     sg.Text('Raw Data Directory Entered', key='rawData', visible=False),
     sg.Text('Invalid File Path Entered', key='invalid', visible=False)]
    ]
    
    
# Compiling all columns into window layout
layout = [
    [sg.Column(filePathCol, key='filePathCol', element_justification='c')],
    [sg.Column(inputCol, justification='center', key='in'), sg.Column(outputCol, visible=False, key='out')],
    [sg.Button('Calculate Kd', key='calculate', disabled=True, disabled_button_color='gray'),
     sg.Button('Report', key='report', visible=False)]
    ]

window = sg.Window("prACTISed", layout, finalize=True, element_justification='c')



######### DEFINE FUNCTIONS ##############

# Display image from file path - for GUI image viewer
def load_image(path,window):
    img = Image.open(path)
    img.thumbnail((350,350))       # (420, 420) is same width as summary  table
    photo_img = ImageTk.PhotoImage(img)
    window['graphImage'].update(data=photo_img)

location = 0


######### EVENT LOOP ##############
while True:
    event, values = window.read() 
    if event == sg.WIN_CLOSED:
        break

# Radiobuttons displaying relevant secondary fields
    # If compensation needed request concentration for normalization
    if event == 'compYes':
        window['normConc'].update(visible=True)
        window['normConc_val'].update(visible=True)
        
    if event == 'compNo':
        window['normConc'].update(visible=False)
        window['normConc_val'].update(visible=False)

    # If manual peak determination, enter peak times
    if event == 'peakM':
        window['manPeaks'].update(visible=True)
        window['manPeaks_val'].update(visible=True)
        window['progPeak'].update(visible=False)
        window['progPeak_val'].update(visible=False)
        
    # If programmatic peak determination, enter concentration to use
    if event == 'peakP':
        window['manPeaks'].update(visible=False)
        window['manPeaks_val'].update(visible=False)
        window['progPeak'].update(visible=True)
        window['progPeak_val'].update(visible=True)

# Validate user input file path
    if event == 'validate':
        window['out'].update(visible=False)
        window['report'].update(visible=False)

        # Confirm input is a string
        check = valStr(values['filePath_val'])
        if check[0] == True:
            filePath = check[1]
        elif check[0] == False:
            sg.popup(check[1])

        # Verify if input path exists
        if not os.path.exists(filePath):
            window['in'].update(visible=False)
            window['calculate'].update(disabled=True)

            window['invalid'].update(visible=True)
            window['work'].update(visible=False)
            window['rawData'].update(visible=False)
            
        # Verify if input path is Excel workbook or directory
        elif os.path.exists(filePath):
            if os.path.isfile(filePath) and filePath.endswith(".xlsx"):
                window['in'].update(visible=False)
                window['calculate'].update(disabled=False)

                window['invalid'].update(visible=False)
                window['work'].update(visible=True)
                window['rawData'].update(visible=False)
                pathType = "workingFile"
                
            elif os.path.isdir(filePath):
                window['in'].update(visible=True)
                window['calculate'].update(disabled=False)

                window['invalid'].update(visible=False)
                window['work'].update(visible=False)
                window['rawData'].update(visible=True)
                pathType = "directory"
   
# Validate user input and calculate Kd     
    if event == 'calculate':

        window['out'].update(visible=False)
        valid = False

        another = [[sg.Text('Validating input fields...', key='loadText')],
                   [sg.ProgressBar(10, orientation='h', size = (20,20), key = 'progressBar')]]
        window2 = sg.Window('prACTISed progress', another, finalize=True)
        if event == sg.WIN_CLOSED:
            break

        # If working file entered, validate contents
        if pathType == "workingFile":
                workingFile = str(values['filePath_val'])

                # Validate workbook & sheets & values
                window2['progressBar'].UpdateBar(1)
                fileResults = validateExcel(workingFile)

                if fileResults[0]==False:
                    errorMessage = genFileErrorMessage(fileResults[1])
                    sg.popup(errorMessage)
                    window2.close()

                elif fileResults[0]:
                    valid = True
 
                    window2['loadText'].update('Reading input Excel file...')
                    window2['progressBar'].UpdateBar(2)
                    data = pd.read_excel(workingFile, engine='openpyxl')
                    inputBook = load_workbook(workingFile, data_only=True)         
                    idealSheet = inputBook["Inputs"]
                    compYN = str(idealSheet.cell(13,2).value)
                
        # If file path is directory validate required user inputs 
        elif pathType == "directory":
            filePath = str(values['filePath_val'])
            window2['progressBar'].UpdateBar(1)
            reqResult = confirmRequired(values)

            if reqResult[0]==False or reqResult[2]==False:
                errorMessage = genErrorMessage(reqResult[1], reqResult[3])
                sg.popup(errorMessage)
                window2.close()

            elif reqResult[0] and reqResult[2]:
                valid = True
                
                # Read in user inputs for fluidics parameters
                propFlow = values['propFlow_val']
                injectFlow = values['injectFlow_val']
                injectTime = float(values['injectTime_val'])
                sepLength = values['sepLength_val']
                sepDiam = values['sepDiam_val']
                injectLength = values['injectLength_val']
                injectDiam = values['injectDiam_val']

                # Read in user inputs for concentrations
                proteinName = str(values['protName_val'])
                ligandName = values['ligName_val']
                ligandConc = float(values['ligConc_val'])

                # Read in user inputs for data analysis parameters
                if window['dataF'].get() == True:
                    dataType = "F"
                elif window['dataMS'].get() == True:
                    dataType = "MS"
                    
                if window['compYes'].get()== True:
                    compYN = "Y"
                    normalConc = float(values['normConc_val'])
                elif window['compNo'].get()== True:
                    compYN = "N"
                    normalConc = None

                windowWidth = values['window_val']

                if values['peakM'] == True:
                    peakDet = "M"
                    manualPeaks = values['manPeaks_val']
                    peakConc=None
                elif values['peakP'] == True:
                    peakDet = "P"
                    manualPeaks = None
                    peakConc = float(values['progPeak_val'])

            
                # Validate contents of raw data directory
                window2['loadText'].update('Validating input directory...')
                window2['progressBar'].UpdateBar(2)
                checkDirect = validateDirectoryContents(filePath, compYN, normalConc, peakDet, peakConc)

                if checkDirect[0] == False:
                    errorMessage = genErrorMessageDirect(checkDirect[1])
                    sg.popup(errorMessage)
                    window.close()
                    window2.close()
                    break
                

                # Check for existing Excel files with same name
                suggName = checkDirect[2]

                exists = checkForOverwriting(suggName)

                if exists[0] == True:
                    if sg.popup_yes_no('A file with the file path %s already exists in this directory. \n \n Would you like to overwrite this file?' % os.path.basename(suggName)) == 'No':
                        workingFilePath = exists[1]

                    else:
                        os.remove(suggName)
                        workingFilePath = (suggName)
                        
                    
                elif exists[0] == False:
                    workingFilePath = suggName

                
                # Prepare working file
                window2['loadText'].update('Preparing working file...')
                window2['progressBar'].UpdateBar(3)
                workingFile = workingfileprep(filePath, workingFilePath, propFlow, injectFlow, injectTime, sepLength, sepDiam,
                                              injectLength, injectDiam, proteinName, ligandName, ligandConc, dataType,
                                              compYN, normalConc, windowWidth, peakDet, manualPeaks, peakConc)

                
        # Unmask signals with compensation procedure if indicated
        if valid:
            
            if compYN == "Y" :
                window2['loadText'].update('Compensating data...')
                window2['progressBar'].UpdateBar(4)
    
                if compensate(workingFile) == False:
                    window2.close()
                    valid = False

                
        if  valid:
            # Analyze signals, update Excel working file and generate subfolder with separagrams and binding isotherm
            window2['loadText'].update('Analyzing data...')
            window2['progressBar'].UpdateBar(5)

            graphPath = dataanalysis(workingFile)
            if graphPath == False:
                window2.close()
                valid = False


        if valid:
            # Load graoh images to display
            window2['loadText'].update('Formatting output data...')
            window2['progressBar'].UpdateBar(9)
            images = natsorted(glob.glob('%s/*.png' % graphPath))
            load_image(images[0],window)
                
            # Read in summary and Kd information from working file to display in GUI output
            df = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="D:J", engine='openpyxl')
            df = df.dropna(how='all')
            headers = df.iloc[0].values.tolist()
            data = df.iloc[1:].values.tolist()
            window['summary'].update(values=data, num_rows=min(10,len(data)))

            df = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="L", engine='openpyxl')
            df = df.dropna(how='all')
            data = df.values.tolist()
            window['Kd'].update(values=data, num_rows=3)
            window2['loadText'].update('prACTISed complete!')
            window2['progressBar'].UpdateBar(10)
            
            # Show output column and report
            window['out'].update(visible=True)
            window['report'].update(visible=True)
            window2.close()
            
        
    # Buttons for image viewer in GUI
    if event == 'fwd':
        if location == len(images)-1:
            location=0
        else:
            location +=1
        load_image(images[location], window)

    if event == 'back':
        if location == 0:
            location=len(images)-1
        else:
             location -=1
        load_image(images[location], window)

    # Functionality for report button to generate, save and open PDF report
    if event == 'report':
        report(workingFile, graphPath) 

window.close()


