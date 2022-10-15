#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# practised_converter_template.py outlines how to create an Excel working file
# which ever dectector is being used with ACTIS
# The workingfileprep function is called by practised.py

# Copyright (C) 2022  Jessica Latimer

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import argparse                                   
import pathlib
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import math
import os
from csv import reader
from natsort import natsorted
    
# Function workingfileprep() to be called by practised.py
# Input parameters passed from GUI
def workingfileprep(inputPath, workingFilePath, propFlow, injectFlow, injectTime, sepLength, sepDiam,
                    injectLength, injectDiam, proteinName, ligandName, ligandConc,
                    dataType, compYN, normalConc, windowWidth, peakDet, manualPeaks, peakConc):

    
    
    # Generate a dictionary d with unitless concentration as key and pandas dataframe as value
    # The pandas dataframe should contain 'raw time' and 'Experiment #' (where # is the run number) columns

    d = {}
    timeRun = {}

    # Reading in files from directory
    for file in os.listdir(inputPath):


##### MODIFY THE FOLLOWING CODE FOR YOUR DESIRED FILE TYPE #####

        # If compensation is required, read in simulated protein profile
        if compYN == "Y" and file.endswith((".txt")) and file.startswith('simulated'):

             simulated = pd.read_csv("%s/%s" % (inputPath,file), sep="\s+", encoding="latin-1", skiprows=1, header=None, keep_default_na=True, na_values=str(0))
             simulated = simulated.fillna(0)
             simulated.columns = ['raw time', 'signal']

             # Normalize simulated protein signals
             maxSim = simulated['signal'].max()
             simulated['signal'] = simulated['signal'].div(maxSim)


        # Read in ASCII raw data files 
        if file.endswith(".asc") and not file.startswith('simulated') and not 'READ' in os.path.splitext(file)[0]:

            # Extract concentration prefix
            molar = file.find("M")
            prefix = file[molar-1]

            # Extract concentration and run number from file name (see naming conventions)
            conc = float(file.partition(prefix)[0])

            name = os.path.splitext(file)[0]
            if name[-1].isdigit()==True: 
                 if name[-2:].isdigit()==True: 
                    runNumber= int(name[-2:])
                 else:
                    runNumber= int(name[-1])
            else:
                runNumber=1


            # Extract the information from ASCII files
            elif file.endswith((".asc"):
                run = pd.read_csv("%s/%s" % (inputPath,file), sep= '\t', encoding="latin-1", keep_default_na=True, na_values=str(0))
                run = run.dropna(how="all")
                run = run.fillna(0)
                run = run.iloc[:,[0,1]]
                run.columns = ["raw time", "Experiment " + str(runNumber)]


                # Create or add experiment to dataframe if it exists
                if conc in d:
                    d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run["Experiment " + str(runNumber)])

                    if runNumber < timeRun[conc]:
                        d[conc].loc[:,'raw time'] = run.iloc[:,0]
                        timeRun[conc] = runNumber
                            
                elif conc not in d:
                    d[conc]=run
                    timeRun[conc] = runNumber


##### MODIFY THE CODE ABOVE FOR YOUR DESIRED FILE TYPE #####                            
    
    # Sort dictionary d alphanumerically 
    for xConc in d:
        orderedExp = natsorted(list(d[xConc].columns))
        orderedExp.remove("raw time")
        orderedExp.insert(0,"raw time")
        d[xConc] = d[xConc].reindex(columns = orderedExp)

    orderedDict = natsorted(d.keys())

    # Optional prefix formatting 
    if prefix == "u":
        prefix = "µ"

    # Create Excel working file
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"

    writer = pd.ExcelWriter(workingFilePath, engine = 'openpyxl')
    writer.book = wb
        
    # Generate Inputs Sheet in Excel working file
    inputDictionary = {"Propogation flow rate":propFlow, "Injection flow rate":injectFlow, "Injection time (s)":injectTime,
                       "Separation capillary length":sepLength, "Separation capillary diameter":sepDiam,
                       "Injection loop length":injectLength, "Injection loop diameter":injectDiam, "Protein name":proteinName,
                       "Ligand name":ligandName, "Number of Concentrations":len(d.keys()),
                       "Initial Ligand concentration [L]0 (%sM)" % prefix: ligandConc, "Type of Data":dataType,
                       "Compensation procedure": compYN, "[P]0 reference for MS normalization (%sM)" % prefix: normalConc,
                       "Window width (%)": windowWidth, "Determination of peak": peakDet, "Manually determined peaks":manualPeaks,
                       "[P]0 to programmaticlly determine peak":peakConc
                       }

    for key in inputDictionary:
        row = list(inputDictionary.keys()).index(key)+1
        ws["A%s" % row] = key
        ws["B%s" % row] = inputDictionary[key]
                   
    for x in range(1,len(d)+1):
        ws["D"+str(x)] = "Protein Conc. #" + str(x)
        ws["E"+str(x)] = "%s %sM" % (orderedDict[x-1], prefix)


   # Add simulated protein profile if compensation required
    if compYN == "Y":
        simulated.to_excel(writer, sheet_name = "P_simulated", index=False)

    # Add sheet for each concentration dataframe
    for y in orderedDict:
        d[y].to_excel(writer, sheet_name = "%s %sM" % (y, prefix), index=False)
            
    # Save and close working file
    writer.save()
    writer.close()
    wb.save(workingFilePath)

    return workingFilePath
