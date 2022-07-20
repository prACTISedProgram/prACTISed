#!/usr/bin/env python3
# -*- coding: utf-8 -*-                 

# ACTIS - Kd Determination Program
# July 12, 2022

#20220708 JL    NOTE: Program no longer generates temporary Excel files
#20220711 JL    NOTE: Verbose statements added, all graphs returned at end of script
#20220712 JL    NOTE: Added: export data to inputted Excel, superimposing separagrams and option to choose concentration to calculate peak signal
#20220720 JL    NOTE: Formatting graphs, allows flexibility to drop a run, and input sheet copied to outputsheet 

# This program extracts ACTIS titration data in a Microsoft Excel file (.xlsx) organized in a particular way*,
# and determines the signal (average peak height within a detection window) for each concentration and a corresponding
# R value for each concentration, then plots a binding isotherm for R vs Protein Concentration and performs
# non-linear curve fitting to calculate and output the Kd value for the experiment

# The program code as shown by default below requires the Microsoft Excel workbook to be organized in the following format:
# - Data for each concentration must be contained in separate worksheets within the Excel file, with each sheet being named in the following format: "# µM".
# - The time intervals are written in column A
# - Row 2 of each worksheet must be the first row containing data
# - Cell A1 is denoted as "raw time"
# - Cells A# are denoted as "Experiment #"
# - The signal measurement for each run is written in each corresponding column of the worksheet

# Testing script execution time
import time
start = time.time()

## Part 1 - Importing the required libraries and sub-libraries required below
import argparse                                   
import pathlib
import sys
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from scipy.optimize import curve_fit

   
## Part 2 - Argument setup and parsing
parser = argparse.ArgumentParser(description = 'prACTISed! This program analyzes ACTIS data and extracts the Kd-value.')
parser.add_argument('inputfile', action = 'store', nargs = 1, type = str)
parser.add_argument('--version', help = 'prints version information', action = 'version', version = 'prACTISed written by Shiv Jain.')
parser.add_argument('-v', '--verbose', help = 'prints detailed output while analyzing', action = 'store_true')

args = vars(parser.parse_args())


## Part 3 - Locating the raw data file and establishing important inputs
fileName = args['inputfile'][0]                         # It's a list but we can only work with one file                           
if not pathlib.Path(fileName).is_file():
        print("Given file '%s' is not a file or does not exist." % fileName)
        exit(-1)

name = pathlib.PurePath(fileName).name

# Set explicitly the engine to use openpyxl, otherwise it might use xlrd, which has removed support for Excel's xlsx format (only supports the old binary format)
data = pd.read_excel(fileName, engine='openpyxl')
inputBook = load_workbook(fileName, data_only=True)         
inputBooknames = inputBook.sheetnames


# Confirm Inputs sheet with correct formatting before preceding
if "Inputs" in inputBook.sheetnames:                                            
    idealSheet = inputBook["Inputs"]
    percentage = int(idealSheet.cell(12,2).value)/100
    numberOfConcs = int(idealSheet.cell(8,2).value)
    injectionTime = int(idealSheet.cell(3,2).value)
    ligandConc = float(idealSheet.cell(9,2).value)
    
    proteinName = str(idealSheet.cell(6,2).value)
    dataType = str(idealSheet.cell(10,2).value)
    peakDet = str(idealSheet.cell(13,2).value)
    
    
elif "Inputs" not in inputBook.sheetnames:  
        sys.exit("Inputfile Formatting Error: The script expects inputdata.xlsx to be in a certain format, see provided idealinputs.xlsx as an example.") 

# Temporary variables
concentration = []
signal = []
stddev = []
relstddev = []
Rvalue = []
Rstddev = []
graphs = []
graphNames = []
forDF = [concentration,signal,stddev,relstddev,Rvalue,Rstddev]
DFnames = ["Conc (µM)","Avg Signal","Std Dev","Rel Std Dev","R value","R Std Dev"]


# If programmatic determination of peak use first run at specified concentration to calculate peak time and time window
if peakDet == "P":
            windowCalcConc = float(idealSheet.cell(15,2).value)

            data = pd.read_excel(fileName, "%.1f µM" % windowCalcConc, engine='openpyxl')
            data = data.dropna(how='all')
            xvalues = data['raw time']
            yvalues = data.iloc[:,1]

            peakSignal = max(yvalues[xvalues>=injectionTime])
            peakIndex = yvalues[yvalues == peakSignal].index
            peakTime = xvalues[peakIndex]

            if args['verbose']:
                    # Prints the peak signal value and the corresponding time
                    print("Peak signal: %.4f a.u. at %.4f seconds" % (peakSignal, peakTime))
                    exit

## Part 4 - Calculating signal information for each concentration and generating separagram graphs
for x in range(1,int(numberOfConcs)+1):         
    conc1 = float(idealSheet.cell(x,5).value)
    numberOfRuns = idealSheet.cell(x,7).value
    
    avgSigConc = 0
    avgSigsRun = []


    # Reading in the whole data frame (= all runs for one particular concentration) and dropping all lines that are blank, i.e. that would produce "NaN"s
    data = pd.read_excel(fileName, "%.1f µM" % conc1, engine='openpyxl')
    data = data.dropna(how='all')
            

    # Calculate background signals for each run
    for col in range(1, numberOfRuns):

        runName = data.columns[col]
        
        if args['verbose']:
            # Prints the concentration and run number
            print("-------- Concentration %.1f µM -------- %s  -------- " % (conc1, runName))
            exit

        xvalues = data['raw time']
        yvalues = data[runName]
        
        # All y-values before the injection time are considered background signal 
        background_yvalues = yvalues[xvalues < injectionTime]
        background_average = np.average(background_yvalues)
        background_stdev = np.std(background_yvalues)

        if args['verbose']:
            # Prints the number of background signals with average and standard deviation
            print("Background signal (first %d values): %.4f±%.4f a.u." % (len(background_yvalues), background_average, background_stdev))
            exit
            

        # Set time window parameters using peak time and determine the average signal within window for each run             
        windowLow = float(peakTime - (percentage * peakTime))
        windowHigh = float(peakTime + (percentage * peakTime))
        windowIndex = xvalues.between(windowLow,windowHigh)
        windowTimes = xvalues[windowIndex]
        windowSignals = yvalues[windowIndex]
        avgSigsRun.append(np.average(windowSignals))
        
        if args['verbose']:
                        # Prints the time window and average signal 
                        print("Time window: %.4f - %.4f seconds. Average signal: %.4f a.u." % (windowLow, windowHigh,np.average(windowSignals)))
                        exit

        # Graph the signal for each run and with time window indicated
        plt.plot(xvalues, yvalues)


    # Appending a figure with all experimental runs for concentration
    plt.xlabel('Propagation time (s)', fontweight='bold')
    if dataType == "MS":
            plt.ylabel('MS intensity (a.u.)', fontweight='bold')
    elif dataType == "F":
            plt.ylabel('Fluorescence (a.u.)', fontweight='bold')
    plt.text(0,peakSignal*1.05, "[%s]0 = %.1f µM" % (proteinName, conc1), fontweight='bold')
    plt.vlines(windowLow, 0, peakSignal*1.05, linestyles='dashed',color='gray')
    plt.vlines(windowHigh, 0, peakSignal*1.05, linestyles='dashed',color='gray')
    plt.savefig("%.1f µM.jpeg" % conc1)    # save separagram graphs
    graphs.append(plt.figure())
    graphNames.append("%.1f µM.jpeg" % conc1)
    plt.clf()


    # Calculating average signal for each concentration, stdev and relative stdev        
    avgSigConc = np.average(avgSigsRun)
    avgSigConc_stdev = np.std(avgSigsRun)
    avgSigConc_relstdev = (avgSigConc_stdev/avgSigConc)*100

    if args['verbose']:
            # Prints the average signal per concentration with standard deviation
            print("Average signal %.1f µM: %.4f±%.4f a.u." % (conc1, avgSigConc, avgSigConc_stdev))
            exit
            
    concentration.append(conc1)
    signal.append(avgSigConc)    
    stddev.append(avgSigConc_stdev)
    relstddev.append(avgSigConc_relstdev)

    
# Graphing separagrams for the first run for every concentration
for x in range(1,int(numberOfConcs)):         
    conc1 = idealSheet.cell(x,5).value

    # Reading in the whole data frame (= all runs for one particular concentration) and dropping all lines that are blank, i.e. that would produce "NaN"s
    data = pd.read_excel(fileName, "%.1f µM" % conc1, engine='openpyxl')
    data = data.dropna(how='all')
    xvalues = data['raw time']
    yvalues = data['Experiment 1']

    plt.plot(xvalues, yvalues, label='%.1f µM' % (conc1))

plt.xlabel('Propagation time (s)', fontweight='bold')
if dataType == "MS":
        plt.ylabel('MS intensity (a.u.)', fontweight='bold')
elif dataType == "F":
        plt.ylabel('Fluorescence (a.u.)', fontweight='bold')
plt.text(0,peakSignal*1.05, "[%s]0" % (proteinName), fontweight='bold')
plt.legend()
plt.vlines(windowLow, 0, peakSignal*1.05, linestyles='dashed',color='gray')
plt.vlines(windowHigh, 0, peakSignal*1.05, linestyles='dashed',color='gray')


# Appending a figure with first experimental run for all concentrations
plt.savefig("allconcentration.jpeg")           # save separagram graph
graphs.append(plt.figure())
plt.clf()


## Part 5 - Calculate R values and standard deviation of R values for each concentration
LowProt_sig = signal[0]
HighProt_sig = signal[numberOfConcs-1]
LowProt_stddev = stddev[0]
HighProt_stdDev = stddev[numberOfConcs-1]

for y in range(0, numberOfConcs):
        conc2 = idealSheet.cell(y+1,5).value
        avgSigConc_R = (signal[y] - HighProt_sig)/ (LowProt_sig - HighProt_sig) 
        Rvalue.append(avgSigConc_R)

        avgSiglConc_Rstddev = ( 1/(LowProt_sig - HighProt_sig) * math.sqrt( (stddev[y]**2) + ((signal[y] - LowProt_sig)/ (LowProt_sig - HighProt_sig) * HighProt_stdDev)**2 +
                                                                       ((HighProt_sig - signal[y])/(LowProt_sig - HighProt_sig) * LowProt_stddev)**2))
        Rstddev.append(avgSiglConc_Rstddev)

        if args['verbose']:
            # R value and standard deviation for each concentration
            print("R value %.1f µM: %.4f±%.4f" % (conc2, avgSigConc_R, avgSiglConc_Rstddev))
            exit

        
if args['verbose']:
            # Summary dataframe of average signal per concentration with standard deviation, relative standard deviation, R value and standard deviation
            df = pd.DataFrame (forDF).transpose()
            df.columns = DFnames
            print(df)
            exit

    
## Part 6 - Plotting the binding isotherm R vs P[0] with curve of best fit
# Plotting data points for each concentration
plt.scatter(concentration, Rvalue, c='white', edgecolor='black', label="R", zorder=10)
plt.errorbar(concentration, Rvalue, yerr = Rstddev, linestyle="none", ecolor = 'black', elinewidth=1, capsize=2, capthick=1, zorder=0)
plt.xscale("log")
              
# Define the Levenberg Marquardt algorithm
def LevenMarqu(x,a):          
    return -((a + x - ligandConc)/(2*ligandConc)) + ((((a + x - ligandConc)/(2*ligandConc))**2) + (a/ligandConc))**(0.5)

# Curve fitting and plotting curve of best fit
popt, pcov = curve_fit(LevenMarqu, concentration, Rvalue)
error = np.sqrt(np.diag(pcov))

xFit = np.arange(0.0, float(conc1), min(concentration))
plt.plot(xFit, LevenMarqu(xFit, popt), linewidth=1.5, color='black', label="Best Fit")
plt.text(0.5, 0.2, 'Kd = %.2f ± \n %.2f μM' % (popt, error), fontweight='bold')
plt.ylabel('R', fontweight='bold')
plt.xlabel('[%s]0 (μM)' % proteinName, fontweight='bold')
plt.xscale("log")
plt.legend()
plt.savefig("bindingisotherm.jpeg")           # save binding isotherm graph
graphs.append(plt.figure())
plt.close()

# Statistics
residuals = Rvalue - LevenMarqu(concentration, *popt)
ss_res = np.sum(residuals**2)
ss_tot = np.sum((Rvalue - np.mean(Rvalue))**2)
r_squared = 1 - (ss_res/ss_tot)

chiSquared = sum((((Rvalue - LevenMarqu(concentration, *popt))**2) / LevenMarqu(concentration, *popt)))

# Returned/printed values
print("Kd: %.4f ± %.4f μM" % (popt,error))
print("R² %.4f" % (r_squared))
print("χ²: %.4f" % (chiSquared))


## Part 7 - Returning summary data and graphs
# Summary dataframe of average signal per concentration with standard deviation, relative standard deviation, R value and standard deviation
df = pd.DataFrame (forDF).transpose()
df.columns = DFnames

# Create new output sheet in input Excel file with summary data and Kd, R² and χ²
writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
writer.book = inputBook
df.to_excel(writer, sheet_name = "Outputs", index=False, startcol=9)        # does not overwrite if a sheet named Outputs already exists
writer.save()
writer.close()

outputSheet = inputBook.worksheets[len(inputBook.sheetnames)-1]

maxr = idealSheet.max_row
maxc = idealSheet.max_column
for r in range(1, maxr+1):
        for c in range (1, maxc+1):
                outputSheet.cell(row=r, column=c).value = idealSheet.cell(row=r, column=c).value

outputSheet["J13"] = "Kd"
outputSheet["K13"] = "%.4f ± %.4f μM" % (popt,error)
outputSheet["J14"] = "R²"
outputSheet["K14"] = "%.4f" % (r_squared)
outputSheet["J15"] = "χ²"
outputSheet["K15"] = "%.4f" % (chiSquared)
inputBook.save(fileName)


# Testing script execution time
end = time.time()
print("Script run time: %.2f seconds" %(end-start))

# Returning all graphs (separagrams and binding isotherm)
plt.show()
