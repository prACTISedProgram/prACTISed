#20220725 JL    NOTE: Switched to PySimpleGUI, added data analysis and working file preparation as functions (compensation & report WIP)
#20220809 JL    NOTE: Added compensation with interpolation from simulated, small adjustments to working file and graph formatting (TODO: report)
#20220912 JL    NOTE: Added pdf report option to GUI (TODO: unicode for Chi character, open pdf)
#20220918 JL    NOTE: Beginning of error trapping, minor refactoring, rearranging GUI, implicit R=[P]0 for programmatic peak determination (TODO: add .dat files to workingfileprep)

import PySimpleGUI as sg

import os
import pathlib
import glob
import sys

from PIL import Image, ImageTk

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from scipy.optimize import curve_fit
from scipy import interpolate
from datetime import date
import time
from natsort import natsorted
from fpdf import FPDF
import webbrowser
#from working import workingfileprep
#from analysis import dataanalysis
#from compensation import compensate
#from pdf import report
#from validate import *

sg.theme('DarkBlue3')

# Building input column for GUI    
inputCol = [
    [sg.Frame('Fluidics Experimental Parameters', 
              [[sg.Text('Propagation flow rate', size=(27,1), key ='propFlow'),
                sg.Input(size=(14,1), key='propFlow_val')],
               [sg.Text('Injection flow rate', size=(27,1),key='injectFlow'),
                sg.Input(size=(14,1), key='injectFlow_val')],
               [sg.Text('Injection time (s) *', size=(27,1), key='injectTime'),
                sg.Input(size=(14,1), key='injectTime_val')],
               [sg.Text('Separation capillary length', size=(27,1), key='sepLength'),
                sg.Input(size=(14,1), key='sepLength_val')],
               [sg.Text('Separation capillary diameter', size=(27,1), key='sepDiam'),
                sg.Input(size=(14,1), key='sepDiam_val')],
               [sg.Text('Injection loop length', size=(27,1), key='injectLength'),
                sg.Input(size=(14,1), key='injectLength_val')],
               [sg.Text('Injection loop diameter', size=(27,1), key='injectDiam'),
                sg.Input(size=(14,1), key='injectDiam_val')]],
              key = 'fluidParam')],


    [sg.Frame('Concentrations', 
              [[sg.Text('Protein name *', size=(27,1), key ='protName'),
                sg.Input(size=(14,1), key='protName_val')],
               [sg.Text('Ligand name', size=(27,1), key='ligName'),
                sg.Input(size=(14,1), key='ligName_val')],
               [sg.Text(u'Initial ligand concentration [L]\u2080 *    \n (same units as [P]\u2080)', key='ligConc'),
                sg.Input(size=(14,1), key='ligConc_val')]],
              key='concFrame')],
    
    [sg.Frame('Data Analyis Parameters',
              [[sg.Text('Type of data *', size=(11,1), key ='dataType'),
                sg.Radio('Fluoresence', 'dataChoice', key ='dataF', enable_events=True),
                sg.Radio('Mass Spec', 'dataChoice', key ='dataMS', enable_events=True)],
               [sg.Text('Compensation procedure *     \n (recommended for MS data)', key ='comp'),
                sg.Radio('Yes', 'compChoice', key ='compYes', enable_events=True),
                sg.Radio('No', 'compChoice', key ='compNo', enable_events=True)],
               [sg.Text(u'[P]\u2080 reference for normalization *', size=(27,1), key='normConc', visible=False),
                sg.Input(size=(14,1), key='normConc_val', visible=False)],
               [sg.Text('Window width (%) *', size=(27,1), key='window'),
                sg.Input(size=(14,1), key='window_val')],
               [sg.Text('Determination of peak *', size=(18,1), key ='peak'),
                sg.Radio('Manual', 'peakChoice', key ='peakM', enable_events=True),
                sg.Radio('Program', 'peakChoice', key ='peakP', enable_events=True)],
               [sg.Text(u'Peaks for concentrations [P]\u2080 * \n (ascending [P]\u2080, separated by commas)', key='manPeaks', visible=False)],
               [sg.Input(size=(43,1), key='manPeaks_val', visible=False)],
               [sg.Text(u'Specify [P]\u2080 used to determine peak *', size=(30,1), key='progPeak', visible=False),
                sg.Input(size=(11,1), key='progPeak_val', visible=False)]],
              key='analysisParam')],
    
    [sg.Text('* Required field', key = 'req'),]
    ]


# Builing output column for GUI        
outputCol= [
    [sg.Table(values=[],headings=['prACTISed output'], key='Kd', hide_vertical_scroll=True, def_col_width=20, auto_size_columns=False)],

    [sg.Table(values=[], headings=['Conc','Avg Signal', 'Std Dev', 'Rel Std Dev', 'R value', 'R Std Dev'], key='summary', def_col_width=10,auto_size_columns=False)],

    [sg.Frame('Graphs',
              [[sg.Image(key='graphImage')],
               [sg.Button('Prev', key='back'),
                sg.Button('Next', key='fwd')]]
               )]           
    ]

# Building file path column for GUI
filePathCol= [
    [sg.Text('File path *', key ='filePath'),
     sg.Input(size=(50,1), key='filePath_val'),
     sg.Button('Validate', key='validate')],
    [sg.Text('Working File Entered', key='work', visible=False),
     sg.Text('Raw Data Directory Entered', key='rawData', visible=False),
     sg.Text('Invalid File Path Entered', key='invalid', visible=False)]
    ]
    
    
# Compiling all columns into window layout
layout = [
    [sg.Column(filePathCol, key='filePathCol', element_justification='c')],
    [sg.Column(inputCol, justification='center', key='in'), sg.Column(outputCol, visible=False, key='out')],
    [sg.Button('Calculate Kd', key='calculate', disabled=True, disabled_button_color='gray'),
     sg.Button('Report', key='report', visible=False)]
    ]

window = sg.Window("prACTISed", layout, finalize=True, element_justification='c')


######### DEFINE FUNCTIONS ##############

# Display image from file path - for GUI image viewer
def load_image(path,window):
    img = Image.open(path)
    img.thumbnail((350,350))       # (420, 420) is same width as summary  table
    photo_img = ImageTk.PhotoImage(img)
    window['graphImage'].update(data=photo_img)

location = 0





# Subfunctions for validating GUI user inputs 
def valFloat (x):       # percentage, injection time, ligandConc, windowConc
    try:
        float(x)
        return [True, float(x)]
    except ValueError:
        return [False, 'Expected a numerical unitless value, please try again']

def valStr (x):    # non-mandatory user inputs, protein name 
    try:
        str(x)
        return [True, str(x)]
    except ValueError:
        return [False, 'Expected a string, please try again']

def valManualTimes (x):
    valStr(x)
    try:
        x = x.split(",")
        for time in x:
            float(x)
        return [True, x]
    except:
        return [False, 'Expected list of numerical times separated by commas, please try again']

### Validating required fields of GUI user inputs (see above for associated sub-functions)      
def confirmRequired (values):
    is_input = True
    values_missing = []
    is_valid = True
    values_invalid = []

    if len(values['injectTime_val'])==0:
        values_missing.append('Injection time')
        is_input = False
    elif len(values['injectTime_val'])>0:
        check = valFloat(values['injectTime_val'])
        if check[0] == False:
            values_invalid.append('Injection time - %s' % check[1])
            is_valid = False

    if len(values['protName_val'])==0:
        values_missing.append('Protein name')
        is_input = False
    elif len(values['protName_val'])>0:
        check = valStr(values['protName_val'])
        if check[0] == False:
            values_invalid.append('Protein name - %s' % check[1])
            is_valid = False

    if len(values['ligConc_val'])==0:
        values_missing.append(u'Initial ligand concentration [L]\u2080')
        is_input = False
    elif len(values['ligConc_val'])>0:
        check = valFloat(values['ligConc_val'])
        if check[0] == False:
            values_invalid.append(u'Initial ligand concentration [L]\u2080 - %s' % check[1])
            is_valid = False

    if not values['dataF'] and not values['dataMS']:
        values_missing.append('Type of data')
        is_input = False

    if not values['compYes'] and not values['compNo']:
        values_missing.append('Compensation procedure')
        is_input = False

    if values['compYes'] and len(values['normConc_val'])==0:
        values_missing.append(u'[P]\u2080 reference for normalization')
        is_input = False
    elif values['compYes'] and len(values['normConc_val'])>0:
        check = valFloat(values['normConc_val'])
        if check[0] == False:
            values_invalid.append(u'[P]\u2080 reference for normalization - %s' % check[1])
            is_valid = False

    if len(values['window_val'])==0:
        values_missing.append('Window width')
        is_input = False
    elif len(values['window_val'])>0:
        check = valFloat(values['window_val'])
        if check[0] == False:
            values_invalid.append('Window width' % check[1])
            is_valid = False

    if not values['peakM'] and not values['peakP']:
        values_missing.append('Determination of peak')
        is_input = False

    if  values['peakM'] and len(values['manPeaks_val'])==0:
        values_missing.append('Peaks for concentrations')
        is_input = False
    elif values['peakM'] and len(values['manPeaks_val'])>0:
        check = valManualTimes(values['manPeaks_val'])
        if check[0] == False:
            values_invalid.append('Peaks for concentrations - %s' % check[1])
            is_valid = False

    if  values['peakP'] and len(values['progPeak_val'])==0:
        values_missing.append(u'[P]\u2080 used to determine peak')
        is_input = False
    elif values['peakP'] and len(values['progPeak_val'])>0:
        check = valFloat(values['progPeak_val'])
        if check[0] == False:
            values_invalid.append(u'[P]\u2080 used to determine peak - %s' % check[1])
            is_valid = False

    result = [is_input, values_missing, is_valid, values_invalid]
    return result

# Generate error message for validation of GUI user inputs
def genErrorMessage (values_missing, values_invalid):
    errorMessage = ''
    for val in values_missing:
        errorMessage += ('\nMissing' + ": " + val)
    for val in values_invalid:
        errorMessage += ('\nInvalid' + ": " + val)
    return errorMessage

###
def valConc (x):
    valStr(x)
    try:
        molar = x.find("M")
        prefix = x[molar-1]
        y = float(x.partition(prefix)[0])
        return [True, '%s %sM' % (y, prefix)]
    except:
        return [False, 'Expected file name to start with concentration with units, please try again']

def validateDirectoryContents (filePath, compYN, normalConc, peakDet, peakConc):
    fileErrors=[]
    filesValid =True
    simData = False
    rawData = False
    normConc = False
    pPeak = False
    
    if len(os.listdir(filePath)) == 0:
        fileErrors.append('Error: Given file directory %s is empty, please try again' % (filePath))
        filesValid = False
    
    for file in os.listdir(filePath):
        if not file.startswith('simulated') and os.path.splitext(file)[1] in (".txt", ".asc"):
            rawData = True
            preCheck = valConc(file)
            if preCheck[0] == False:
                fileErrors.append('Error: File name %s does not follow prefix conventions - %s' % (file, preCheck[1]))
                filesValid = False
            if preCheck[0] == True and peakDet== "P":
                molar = file.find("M")
                prefix = file[molar-1]
                if float(file.partition(prefix)[0]) == peakConc:
                    pPeak = True
                if type(normalConc)==float:
                    if float(file.partition(prefix)[0]) == normalConc:
                        normConc = True
            
            endCheck = os.path.splitext(file)[0]
            if not endCheck[-1].isnumeric:
                fileErrors.append('Error: File name %s does not follow suffix conventions - Expected file name to end with run number,  please try again' % (file))
                filesValid = False
                                  
        if file.startswith('simulated') and os.path.splitext(file)[1] == '.txt':
            simData = True

    if compYN =="Y" and simData==False:
            fileErrors.append('Error: No simulated protein profile .txt file beginning with simulated found in file directory %s , please try again' % (filePath))
            filesValid = False
    if compYN =="Y" and normConc==False:
            fileErrors.append(u'Error: No files found for indicated [P]\u2080 = %s used for normalization' % (normalConc))
            filesValid = False
        
    if rawData == False:
        fileErrors.append('Error: No .asc or .txt raw data files found in %s' % (filePath))
        filesValid = False

    if peakDet=="P" and pPeak == False:
        fileErrors.append(u'Error: No files found for indicated [P]\u2080 = %s used to programmaticlly determine peak' % (peakConc))
        filesValid = False
        
    return [filesValid, fileErrors]
                
# Generate error message for validating directory to raw data files (& simulated protein profile)
def genErrorMessageDirect (fileErrors):
    errorMessage = ''
    for val in fileErrors:
        errorMessage += ('\n'+ val)
    return errorMessage


### Check if file with same name exists, check if user wants to overwrite
def checkForOverwriting (suggestedFilePath):
    exists = False
    exists_Name = ''
    if os.path.exists(suggestedFilePath)== False:
        return [exists, suggestedFilePath]

    elif os.path.exists(suggestedFilePath)== True:
        exists = True
        exists_Name = suggestedFilePath
        for duplicate in range(2,9):
            withoutExt = os.path.splitext(suggestedFilePath)[0]
            temp = "%s_%d.xlsx" % (withoutExt, duplicate)

            if os.path.exists(temp)== False:
                suggestedFilePath=temp
                return [exists, suggestedFilePath]
                break


### Subfunctions for validating Excel workbook
def valDataType (x):
    valStr(x)
    if x == "F" or x == 'f' or "F" in x or 'f' in x:
        return [True, "F"]
    elif x == "MS" or x == 'ms' or "MS" in x or 'ms' in x:
        return [True, "MS"]
    else:
        return [False, 'Expected F or MS, please try again']


def valComp (x):
    valStr(x)
    if x == "Y" or x == 'y' or "Y" in x or 'y' in x:
        return [True, "Y"]
    elif x == "N" or x == 'n' or "N" in x or ('n' in x and 'en' not in x):
        return [True, "N"]
    elif x == "Compensated" or x == "compensated" or "ompensate" in x:
        return [True, "Compensated"]
    else:
        return [False, 'Expected Y, N or Compensated, please try again']
    

def valPeakDetermination (x):
    valStr(x)
    if x == "P" or x == 'p' or "P" in x or 'p' in x:
        return [True, "P"]
    elif x == "M" or x == 'm' or "M" in x or 'm' in x:
        return [True, "M"]
    else:
        return [False, 'Expected P or M, please try again']
    

def valManualTimes (x, numConcs, injectTime):
    valStr(x)
    try:
        times_valid = True
        y = x.split(",")
        if len(y) != numConcs:
            return [False, 'Error: Manually determined peaks - Expected %s times,  %s inputted' % (numConcs, len(y))]
        for time in y:
            if time < injectTime:
                times_valid = False
                return [False, 'Error: Manually determined peaks - Expected times greater than injection time %s s' % injectTime]
                break
        if len(y)==numConcs and times_valid==True:
                return [True, str(x)]
    except:
        return [False, 'Expected list of times separated by commas, please try again']
    

def valConc (x):
    valStr(x)
    try:
        molar = x.find("M")
        prefix = x[molar-1]
        y = float(x.partition(prefix)[0])
        return [True, '%s %sM' % (y, prefix)]
    except:
        return [False, 'Expected concentration with units, please try again']
    
### Validating an Excel workbook for expected sheets and mandatory fields (see above for associated sub-functions)
def validateExcel (filePath):
    
    # Validate file path is for an Excel workbook
    inputErrors = []
    is_valid = True
    if not os.path.exists(filePath):
        inputErrors.append('Error: Path %s does not exist' % filePath)
        is_valid=False
        
    if not os.path.isfile(filePath):
         inputErrors.append('Error: %s is not a valid file path' % filePath)
         is_valid=False
        
    if not filePath.endswith(".xlsx"):
        inputErrors.append('Error: %s is not a valid Excel file' % filePath)
        is_valid=False
         
    if os.path.isfile(filePath) and filePath.endswith(".xlsx"):
        data = pd.read_excel(filePath, engine='openpyxl')
        inputBook = load_workbook(filePath, data_only=True)
        sheets = inputBook.sheetnames

        # TODO: Confirm Excel workbook is not open, confirm permissions for reading/writing Excel workbook

        # Verify there is a sheet named Inputs
        if 'Inputs' not in sheets:
            return "Error: No sheet named Inputs found in Excel workbook %s" % filePath

        # If Inputs sheet exists, confirm all mandatory inputs
        elif 'Inputs' in sheets:
            idealSheet = inputBook["Inputs"]
            
            windowCheck = valFloat(idealSheet.cell(15,2).value)
            if windowCheck[0]!=True:
                inputErrors.append('Error: Window width - %s' % windowCheck[1])
                is_valid = False

            numConcCheck = valFloat(idealSheet.cell(10,2).value)
            if numConcCheck[0]!=True:
                inputErrors.append('Error: Number of Concentrations - %s' % numConcCheck[1])
                is_valid = False
            
            injectTimeCheck = valFloat(idealSheet.cell(3,2).value)
            if injectTimeCheck[0]!=True:
                inputErrors.append('Error: Injection time - %s' % injectTimeCheck[1])
                is_valid = False
                            
            ligandConcCheck = valFloat(idealSheet.cell(11,2).value)
            if ligandConcCheck[0]!=True:
                inputErrors.append('Error: Initial Ligand concentration - %s' % ligandConcCheck[1])
                is_valid = False
                                   
            proteinNameCheck = valStr(idealSheet.cell(8,2).value)
            if proteinNameCheck[0]!=True:
                inputErrors.append('Error: Protein name - %s' % proteinNameCheck[1])
                is_valid = False
                                   
            # Validate data type
            dataTypeCheck = valStr(idealSheet.cell(12,2).value)
            if dataTypeCheck[0]!=True:
                inputErrors.append('Error: Data Type - %s' % dataTypeCheck[1])
                is_valid = False
            elif dataTypeCheck[0]==True:
                secondCheck = valDataType(dataTypeCheck[1])
                if secondCheck[0] == True:
                    idealSheet['B12'] = secondCheck[1]
                elif secondCheck[0] == False:
                    inputErrors.append('Error: Data Type - %s' % secondCheck[1])
                    is_valid = False

            # Validate compensation specification
            compYNCheck = valStr(idealSheet.cell(13,2).value)
            if compYNCheck[0]!=True:
                inputErrors.append('Error: Compensation procedure - %s' % compYNCheck[1])
                is_valid = False
            elif compYNCheck[0]==True:
                secondCheck = valComp(compYNCheck[1])
                if secondCheck[0] == True:
                    idealSheet['B13'] = secondCheck[1]

                    if secondCheck[1] == "Y":
                        normalConcCheck = valFloat(idealSheet.cell(14,2).value)
                        if normalConcCheck[0] == False:
                            inputErrors.append('Error: Compensation procedure - %s' % secondCheck[1])
                            is_valid = False
                            
                        if 'P_simulated' not in sheets:
                            inputErrors.append("Error: No sheet named P_simulated found in Excel workbook %s" % filePath)
                            is_valid = False          
                            
                elif secondCheck[0] == False:
                    inputErrors.append('Error: Compensation procedure - %s' % secondCheck[1])
                    is_valid = False

            # Validate peak determination method and dependent manual or programmatic times
            peakDetCheck = valStr(idealSheet.cell(16,2).value)
            if peakDetCheck[0]!=True:
                inputErrors.append('Error: Determination of peak - %s' % dataTypeCheck[1])
                is_valid = False
            elif peakDetCheck[0]==True:
                secondCheck = valPeakDetermination(peakDetCheck[1])
                if secondCheck[0] == False:
                    inputErrors.append('Error: Determination of peak - %s' % secondCheck[1])
                    is_valid = False
                elif secondCheck[0] == True:
                    idealSheet['B16'] = secondCheck[1]

                    if secondCheck[1] == "M":
                        if numConcCheck[0]==True and injectTimeCheck[0]==True:
                            manualTimesCheck = valManualTimes(idealSheet.cell(17,2).value, numConcCheck[1], injectTimeCheck[1])
                            if manualTimesCheck[0] == False:
                                inputErrors.append('Error: Manually determined peaks - %s' % secondCheck[1])
                                is_valid = False
                                
                    elif secondCheck[1] == "P":
                        pPeakCheck = valFloat(idealSheet.cell(18,2).value)
                        if pPeakCheck[0]==False:
                            inputErrors.append(u'Error: [P]\u2080 used to determine peak - %s' % pPeakCheck[1])
                            is_valid = False

            if numConcCheck[0]==True:
                for  x in range(1,int(numConcCheck[1])+1):
                    concCheck = valConc(idealSheet.cell(x,5).value)
                    if concCheck[0]==False:
                        inputErrors.append('Error: Listed concentrations - %s' % concCheck[1])
                        is_valid = False
                    elif concCheck[0]==True:
                        idealSheet['%s5' % get_column_letter(x)] = secondCheck[1]
                        if concCheck[1] not in sheets:
                            inputErrors.append('Error: No sheet named %s found in Excel workbook %s' % (concCheck[1],filePath))
                            is_valid = False
                        
                        elif concCheck[1] in sheets:
                            df = pd.read_excel(filePath, sheet_name=concCheck[1], engine='openpyxl')
                            df = df.dropna(how='all')
                
                            if df.columns[0] != 'raw time':
                                inputErrors.append('Error: No raw time column found in %s sheet of Excel workbook %s' % (concCheck[1],filePath))
                                is_valid = False
                                                          
    return [is_valid, inputErrors]

# Generate error message for validation of Excel workbook                                    
def genFileErrorMessage(inputErrors):
    errorMessage = ''
    for val in inputErrors:
        errorMessage += ('\n' + val)
    return errorMessage



def workingfileprep(inputPath, workingFilePath, propFlow, injectFlow, injectTime, sepLength, sepDiam,
                    injectLength, injectDiam, proteinName, ligandName, ligandConc,
                    dataType, compYN, normalConc, windowWidth, peakDet, manualPeaks, peakConc):

    
    ## Part 1 - Importing the required libraries and sub-libraries required below
    import argparse                                   
    import pathlib
    import sys
    import pandas as pd
    import matplotlib.pyplot as plt
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils import FORMULAE
    from openpyxl.utils import get_column_letter, column_index_from_string
    import math
    from scipy.optimize import curve_fit

    import os
    from csv import reader
    from natsort import natsorted
    from datetime import date

    # Testing script execution time
    import time
    start = time.time()


    d = {}
    prefix=""

    def isfloat(num):
        try:
            float(num)
            return True
        except ValueError:
            return False     

    ### Reading in files from directory ###
    for file in os.listdir(inputPath):

        ## Read in simulated protein profile if compensation procedure was selected
        if compYN == "Y" and file.endswith((".txt")) and file.startswith('simulated'):

             # Extract the preamble information
             preamble = []
             with open("%s/%s" %(inputPath, file), "r", encoding="latin-1") as fileCheck:
                 csv_reader = reader(fileCheck, delimiter=" ")
         
                 for row in csv_reader:
                     if row[0].isalpha() == True or isfloat(row[0])==False:
                         preamble.append(row)
                           
                     elif row[0].isalpha() == False or isfloat(row[0])==True:
                         break

             # Read in simulated protein profile if compensation procedure was selected
             simulated = pd.read_csv("%s/%s" % (inputPath,file), sep="\s+", encoding="latin-1", skiprows=len(preamble), header=None, keep_default_na=True, na_values=str(0))
             simulated = simulated.fillna(0)
             simulated.columns = ['raw time', 'signal']

             # Normalize simulated protein signals
             maxSim = simulated['signal'].max()
             simulated['signal'] = simulated['signal'].div(maxSim)


        ## Read in raw data files 
        if file.endswith((".txt", ".asc")) and not file.startswith('simulated'):

            # Extract concentration prefix
            molar = file.find("M")
            prefix = file[molar-1]

            # Extract concentration and run number from file name (see naming conventions)
            conc = float(file.partition(prefix)[0])

            name = os.path.splitext(file)[0]
            if name[-1].isdigit()==True: 
                 if name[-2:].isdigit()==True: 
                    runNumber= int(name[-2:])
                 else:
                    runNumber= int(name[-1])
            elif name[-1] == ' ':
                runNumber=1

            # Extract the preamble information
            preamble = []
            with open("%s/%s" %(inputPath, file), "r", encoding="latin-1") as fileCheck:
                csv_reader = reader(fileCheck, delimiter="\t")
                for row in csv_reader:
                       
                    if row[0].isalpha() == True or isfloat(row[0])==False:
                        preamble.append(row)
                           
                    elif row[0].isalpha() == False or isfloat(row[0])==True:
                        break

            # If no multiplier extract time and signal
            if len(preamble) <= 1:
                run = pd.read_csv("%s/%s" % (inputPath,file), sep="\t", encoding="latin-1", keep_default_na=True, na_values=str(0))
                run = run.dropna(how="all")
                run = run.fillna(0)
                run = run.iloc[:,[0,1]]
                run.columns = ["raw time", "Experiment " + str(runNumber)]
                run.iloc[:,0] = run.iloc[:,0].mul(60)


                # Create or add experiment to dataframe if it exists
                if conc in d:
                    d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run["Experiment " + str(runNumber)])

                elif conc not in d:
                    d[conc]=run

            # If multiplier needed extract signals, signal multipler, and iterate over signals
            elif len(preamble) > 1:
                signalMult_line = list(filter(lambda x: "Y Axis Multiplier:" in x[0], preamble))
                signalMult = float(signalMult_line[0][1])

                run = pd.read_csv("%s/%s" %(inputPath, file), sep="\t", encoding="latin-1", skiprows=len(preamble), header=None, keep_default_na=True, na_values=str(0))
                run = run.dropna(how="all")
                run = run.fillna(0)
                run.columns = ["Experiment" + str(runNumber)]
                run.loc[:"Experiment" + str(runNumber)] *= signalMult

                # Create or add experiment to dataframe if it exists
                if conc in d:
                    d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run)

                elif conc not in d:
                    # Reconstruct raw time from sampling rate
                    hz_line = list(filter(lambda x: "Sampling Rate:" in x[0], preamble))
                    hz = float(hz_line[0][1])
                    second_Gap = 1/hz

                    rawTime = [0]

                    for x in range(0,len(run)-1):
                        rawTime.append(rawTime[x]+second_Gap)

                    run.insert(0, "raw time", rawTime)
                    d[conc]=run
                        
    for xConc in d:
        orderedExp = natsorted(list(d[xConc].columns))
        orderedExp.remove("raw time")
        orderedExp.insert(0,"raw time")
        d[xConc] = d[xConc].reindex(columns = orderedExp)
            
    orderedDict = natsorted(d.keys())

    if prefix == "u":
        prefix = "µ"

    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"

    writer = pd.ExcelWriter(workingFilePath, engine = 'openpyxl')
    writer.book = wb
        
    # Generate Inputs Sheet
    inputDictionary = {"Propogation flow rate":propFlow, "Injection flow rate":injectFlow, "Injection time (s)":injectTime,
                       "Separation capillary length":sepLength, "Separation capillary diameter":sepDiam,
                       "Injection loop length":injectLength, "Injection loop diameter":injectDiam, "Protein name":proteinName,
                       "Ligand name":ligandName, "Number of Concentrations":len(d.keys()),
                       "Initial Ligand concentration [L]0 (%sM)" % prefix: ligandConc, "Type of Data":dataType,
                       "Compensation procedure": compYN, "[P]0 reference for MS normalization (%sM)" % prefix: normalConc,
                       "Window width (%)": windowWidth, "Determination of peak": peakDet, "Manually determined peaks":manualPeaks,
                       "[P]0 to programmaticlly determine peak":peakConc
                       }

    for key in inputDictionary:
        row = list(inputDictionary.keys()).index(key)+1
        ws["A%s" % row] = key
        ws["B%s" % row] = inputDictionary[key]
                   
    for x in range(1,len(d)+1):
        ws["D"+str(x)] = "Protein Conc. #" + str(x)
        ws["E"+str(x)] = "%s %sM" % (orderedDict[x-1], prefix)


   # Add simulated protein profile if compensation required
    if compYN == "Y":
        simulated.to_excel(writer, sheet_name = "P_simulated", index=False)

    # Add sheet for each concentration dataframe
    for y in orderedDict:
        d[y].to_excel(writer, sheet_name = "%s %sM" % (y, prefix), index=False)
            
    writer.save()
    writer.close()
    wb.save(workingFilePath)


    # Testing script execution time
    end = time.time()
    print("Script run time: %.2f seconds" %(end-start))

    return workingFilePath



def compensate (fileName):
#fileName = "/Users/jess/Documents/practised/ALP.xlsx"

        d ={}

        data = pd.read_excel(fileName, engine='openpyxl')
        inputBook = load_workbook(fileName, data_only=True)
        idealSheet = inputBook["Inputs"]

        injectTime = idealSheet.cell(3,2).value
        numberOfConcs = idealSheet.cell(10,2).value
        normalConc = float(idealSheet.cell(14,2).value)
        unit = idealSheet.cell(1,5).value.partition(" ")[2]
        normalConc = "%s %s" % (normalConc, unit)
        compYN = str(idealSheet.cell(13,2).value)

        if compYN == 'Y':
        
                # Get dimensionless simulated separagram of pure protein, S̃p, and interpolate signals
                simulated = pd.read_excel(fileName, sheet_name='P_simulated', engine='openpyxl')

                # Interpolate signals from simulated protein profile
                timeSim = simulated['raw time']
                sigSim = simulated['signal']
                interp = interpolate.splrep(timeSim, sigSim)

                # Isolate interrpolated signals at for times in raw data files
                rawData = pd.read_excel(fileName, sheet_name=normalConc, engine='openpyxl')
                rawData = rawData.dropna(how='all')
                rawTime = rawData['raw time']
                interpSigs = interpolate.splev(rawTime, interp, der=0)

                # Normalize signals, remove negative values and generate data frame
                interpSigs = pd.Series(interpSigs, name='signal')
                interpSigs = interpSigs.div(max(interpSigs))
                interpSigs[interpSigs < 0] = 0
                simulatedSigs = pd.concat([rawTime,interpSigs], axis=1)

                # Get integrated signal of normalization concentration
                integratedNorm = []

                # Isolate first run of concentration used to normalize
                rawSignal = pd.read_excel(fileName, sheet_name=normalConc, engine='openpyxl')
                rawSignal = rawSignal.dropna(how='all')
                rawSignal = rawSignal.iloc[:,:2]

                time = rawSignal['raw time']
                sig = rawSignal[rawSignal.columns[1]]
                colName = str(rawSignal.columns[1])

                # Average the background signals for first 5 secs, subtract from all signals
                background = sig[time < injectTime].mean()
                rawSignal.iloc[:,1] = rawSignal.iloc[:,1].sub(background)

                # Multiply raw signal by Sp
                rawSignal.iloc[:,1] = rawSignal.iloc[:,1].mul(simulatedSigs.iloc[:,1])

                # Normalize and remove negative signals
                rawSignal.iloc[:,1] = rawSignal.iloc[:,1].mul(1)
                rawSignal.loc[(rawSignal[colName] < 0), colName] = 0
                normArea = rawSignal.iloc[:,1].sum()

                # Add concentration to dictionary
                d[normalConc]= rawSignal
                        

                # Apply Sp and normalization to all other concentrations and runs
                for x in range(1,int(numberOfConcs)+1):
                    
                        conc1 = str(idealSheet.cell(x,5).value)

                        # Read in all data for concentration
                        rawSignal = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
                        rawSignal = rawSignal.dropna(how='all')

                        time = rawSignal['raw time']
                        numberOfRuns = len(rawSignal.columns)

                        for run in range(1, int(numberOfRuns)):

                                # Skip normalization concentration run 1
                                if conc1.partition(" ")[0] != "%s" % normalConc or run != 1:
                                
                                        sig = rawSignal[rawSignal.columns[run]]

                                        # Average the background signals for frist 5 secs, subtract from all signals
                                        background = sig[time < injectTime].mean()
                                        sig = sig.sub(background)

                                        # Multiply raw signal by Sp
                                        sig = sig.mul(simulatedSigs.iloc[:,1])

                                        # Normalize
                                        sigArea = sig.sum()
                                        sig = sig.mul(normArea)
                                        sig = sig.div(sigArea)
                                        sig[sig < 0] = 0

                                        # Add concentration to dictionary or append run to exisiting entry
                                        if run == 1:
                                                sig = pd.DataFrame(sig, columns=[rawSignal.columns[run]])
                                                sig.insert(0, "raw time", time)
                                                d[conc1]= sig

                                        else:
                                                d[conc1].insert(len(d[conc1].columns), rawSignal.columns[run], sig)
                

                writer = pd.ExcelWriter(fileName, engine='openpyxl')
                writer.book = inputBook

                writer.sheets = dict((ws.title, ws) for ws in inputBook.worksheets)

                for y in d.keys():
                        d[y].to_excel(writer, sheet_name = y, index=False)

                simulatedSigs.to_excel(writer, sheet_name = 'P_simulated', index=False)

                idealSheet["B13"] = "Compensated"

                writer.save()
                writer.close()
                inputBook.save(fileName)




def dataanalysis(fileName):
        
        # Testing script execution time
        import time
        start = time.time()

        ## Part 1 - Importing the required libraries and sub-libraries required below
        import argparse                                   
        import pathlib
        import sys
        import pandas as pd
        import matplotlib.pyplot as plt
        import numpy as np
        from openpyxl import load_workbook
        from openpyxl import Workbook
        from openpyxl.utils import FORMULAE
        from openpyxl.utils import get_column_letter, column_index_from_string
        import math
        from scipy.optimize import curve_fit
        from datetime import date
        import os


        ## Part 3 - Locating the raw data file and establishing important inputs                         
        if not pathlib.Path(fileName).is_file():
                print("Given file '%s' is not a file or does not exist." % fileName)
                exit(-1)

        name = pathlib.PurePath(fileName).name

        # Set explicitly the engine to use openpyxl, otherwise it might use xlrd, which has removed support for Excel's xlsx format (only supports the old binary format)
        data = pd.read_excel(fileName, engine='openpyxl')
        inputBook = load_workbook(fileName, data_only=True)         
        inputBooknames = inputBook.sheetnames


        # Confirm Inputs sheet with correct formatting before preceding
        if "Inputs" in inputBook.sheetnames:                                            
            idealSheet = inputBook["Inputs"]
            percentage = float(idealSheet.cell(15,2).value)/100
            numberOfConcs = int(idealSheet.cell(10,2).value)
            injectionTime = float(idealSheet.cell(3,2).value)
            ligandConc = float(idealSheet.cell(11,2).value)
            
            proteinName = str(idealSheet.cell(8,2).value)
            dataType = str(idealSheet.cell(12,2).value)
            peakDet = str(idealSheet.cell(16,2).value)

            
            subdirect = "%s_graphs" % os.path.splitext(fileName)[0]
            if os.path.exists(subdirect)== False:
                    os.mkdir(subdirect)
                    
            elif os.path.exists(subdirect)== True:
                    for duplicate in range(2,9):
                            temp = "%s_%d" % (subdirect, duplicate)
                            
                            if os.path.exists(temp)== False:
                                    subdirect=temp
                                    os.mkdir(subdirect)
                                    break
            
        elif "Inputs" not in inputBook.sheetnames:  
                sys.exit("Input file Formatting Error: The script expects inputdata.xlsx to be in a certain format, see provided idealinputs.xlsx as an example.") 

        # Temporary variables
        concentration = []
        signal = []
        stddev = []
        relstddev = []
        Rvalue = []
        Rstddev = []
        graphs = []
        graphNames = []
        forDF = [concentration,signal,stddev,relstddev,Rvalue,Rstddev]
        DFnames = ["Conc","Avg Signal","Std Dev","Rel Std Dev","R value","R Std Dev"]

        # If programmatic determination of peak use first run at specified concentration to calculate peak time and time window
        if peakDet == "P":
                windowCalcConc = float(idealSheet.cell(18,2).value)
                windowCalcConcS = "%s %s" % (windowCalcConc, idealSheet.cell(1,5).value.partition(" ")[2])

                data = pd.read_excel(fileName, sheet_name=windowCalcConcS, engine='openpyxl')
                data = data.dropna(how='all')
                xvalues = data['raw time']
                yvalues = data.iloc[:,1]

                peakSignal = max(yvalues[xvalues>=injectionTime])
                peakIndex = yvalues[yvalues == peakSignal].index
                peakTime = xvalues[peakIndex]

        # Determine absolute max signal value and max number of runs (used to set graph parameters)
        maxSig = 0
        maxRuns = 0
        for x in range(1,int(numberOfConcs)+1):         
                conc1 = idealSheet.cell(x,5).value
                data = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
                data = data.dropna(how='all')
                data = data.iloc[:,1:]
                currentMaxRuns = len(data.columns)
                colMax = data.max()
                currentMax = colMax.max()

                if currentMax > maxSig:
                        maxSig = currentMax
                if currentMaxRuns > maxRuns:
                    maxRuns = currentMaxRuns

                
        ## Part 4 - Calculating signal information for each concentration and generating separagram graphs
        for x in range(1,int(numberOfConcs)+1):         
            conc1 = idealSheet.cell(x,5).value
            
            avgSigConc = 0
            avgSigsRun = []


            # Reading in the whole data frame (= all runs for one particular concentration) and dropping all lines that are blank, i.e. that would produce "NaN"s
            data = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
            data = data.dropna(how='all')
            numberOfRuns = len(data.columns)
            len(data.columns)-1

            # Calculate background signals for each run
            for col in range(1, numberOfRuns):

                runName = data.columns[col]

                xvalues = data['raw time']
                yvalues = data[runName]
                minTime = xvalues[0]
                
                # All y-values before the injection time are considered background signal 
                background_yvalues = yvalues[xvalues < injectionTime]
                background_average = np.average(background_yvalues)
                background_stdev = np.std(background_yvalues)


                # If manual determination, extract the manually set peak time for given concentration
                if peakDet == "M":   
                    manualTimes = idealSheet.cell(17,2).value.split(",")
                    peakIndex = xvalues.searchsorted(float(manualTimes[x-1]), side='left')
                    peakTime = xvalues[peakIndex-1]

                # Set time window parameters using peak time and determine the average signal within window for each run             
                windowLow = float(peakTime - (percentage * peakTime))
                windowHigh = float(peakTime + (percentage * peakTime))
                windowIndex = xvalues.between(windowLow,windowHigh)
                windowTimes = xvalues[windowIndex]
                windowSignals = yvalues[windowIndex]
                avgSigsRun.append(np.average(windowSignals))
          

                # Graph the signal for each run and with time window indicated
                plt.plot(xvalues, yvalues)
                

            # Appending a figure with all experimental runs for concentration
            plt.xlabel('Propagation time (s)', fontweight='bold')
            if dataType == "MS":
                    plt.ylabel('MS intensity (a.u.)', fontweight='bold')
            elif dataType == "F":
                    plt.ylabel('Fluorescence (a.u.)', fontweight='bold')
            plt.text(minTime,maxSig*1.05, r"[%s]$\mathbf{_0}$ = %s" % (proteinName, conc1), fontweight='bold')
            plt.vlines(windowLow, 0, maxSig*1.05, linestyles='dashed',color='gray')
            plt.vlines(windowHigh, 0, maxSig*1.05, linestyles='dashed',color='gray')
            plt.savefig("%s/%s.png" % (subdirect, conc1))
            graphs.append(plt.figure())
            plt.clf()

            
            # Calculating average signal for each concentration, stdev and relative stdev        
            if peakDet == "P" and float(conc1.partition(" ")[0]) >= windowCalcConc:
                avgSigConc = np.average(avgSigsRun)
                avgSigConc_stdev = np.std(avgSigsRun)
                avgSigConc_relstdev = (avgSigConc_stdev/avgSigConc)*100

                        
                concentration.append(conc1)
                signal.append(avgSigConc)    
                stddev.append(avgSigConc_stdev)
                relstddev.append(avgSigConc_relstdev)

            
        # Generate legend for run colors
        #fig = plt.figure()
        #figLeg = plt.figure(figsize=(2, 1.25))
        #labels = []
        #lines = []
        #for r in range (1,maxRuns):
                #ax = fig.add_subplot(111)
                #lines.append(ax.plot([1],[2]))
                #labels.append('Run %s' % r)
                #lines.append((r, sin(
        #plt.legend(labels, loc='center')
        #plt.savefig("%s/lengendM.png" % subdirect)
        #plt.clf()

        # Graphing separagrams for the first run for every concentration
        for x in range(1,int(numberOfConcs),2):         
            conc1 = idealSheet.cell(x,5).value

            # Reading in the whole data frame (= all runs for one particular concentration) and dropping all lines that are blank, i.e. that would produce "NaN"s
            data = pd.read_excel(fileName, sheet_name=conc1, engine='openpyxl')
            data = data.dropna(how='all')
            xvalues = data['raw time']
            yvalues = data.iloc[:, 1]

            p = plt.plot(xvalues, yvalues, label= '%s' % conc1)   

        plt.xlabel('Propagation time (s)', fontweight='bold')
        if dataType == "MS":
                plt.ylabel('MS intensity (a.u.)', fontweight='bold')
        elif dataType == "F":
                plt.ylabel('Fluorescence (a.u.)', fontweight='bold')
        plt.text(minTime, maxSig*1.05, r"[%s]$\mathbf{_0}$"  % (proteinName), fontweight='bold')
        plt.legend()
        if peakDet == "P":
                plt.vlines(windowLow, 0, maxSig*1.05, linestyles='dashed',color='gray')
                plt.vlines(windowHigh, 0, maxSig*1.05, linestyles='dashed',color='gray')
        elif peakDet == "M":
                plt.vlines(0, maxSig, maxSig*1.05, linestyles='dashed',color='white')

        # Appending a figure with first experimental run for all concentrations
        plt.savefig("%s/allconcentration.png" % subdirect)           # save separagram graph
        graphs.append(plt.figure())
        plt.clf()


        ## Part 5 - Calculate R values and standard deviation of R values for each concentration 
        LowProt_sig = signal[0]
        HighProt_sig = signal[len(concentration)-1]
        LowProt_stddev = stddev[0]
        HighProt_stdDev = stddev[len(concentration)-1]

        for y in range(0, len(concentration)-1):
                conc2 = idealSheet.cell(y+1,5).value
                avgSigConc_R = (signal[y] - HighProt_sig)/ (LowProt_sig - HighProt_sig) 
                Rvalue.append(avgSigConc_R)

                avgSiglConc_Rstddev = ( 1/(LowProt_sig - HighProt_sig) * math.sqrt( (stddev[y]**2) + ((signal[y] - LowProt_sig)/ (LowProt_sig - HighProt_sig) * HighProt_stdDev)**2 +
                                                                               ((HighProt_sig - signal[y])/(LowProt_sig - HighProt_sig) * LowProt_stddev)**2))
                Rstddev.append(avgSiglConc_Rstddev)
            
        ## Part 6 - Plotting the binding isotherm R vs P[0] with curve of best fit
        # Convert concentration strings to floats
        concs = []
        for element in range(0, len(concentration)-1):
                num = concentration[element].partition(" ")[0]
                concs.append(float(num))
        unit = concentration[0].partition(" ")[2]

        # Plotting data points for each concentration
        plt.scatter(concs, Rvalue, c='white', edgecolor='black', label="R", zorder=10)
        plt.errorbar(concs, Rvalue, yerr = Rstddev, linestyle="none", ecolor = 'black', elinewidth=1, capsize=2, capthick=1, zorder=0)
        plt.xscale("log")
                      
        # Define the Levenberg Marquardt algorithm
        def LevenMarqu(x,a):          
            return -((a + x - ligandConc)/(2*ligandConc)) + ((((a + x - ligandConc)/(2*ligandConc))**2) + (a/ligandConc))**(0.5)

        # Curve fitting and plotting curve of best fit    
        popt, pcov = curve_fit(LevenMarqu, concs, Rvalue)
        error = np.sqrt(np.diag(pcov))

        step=0
        if concs[step] == float(0):
                step = 1

        xFit = np.arange(0.0, max(concs), concs[step])
        plt.plot(xFit, LevenMarqu(xFit, popt), linewidth=1.5, color='black', label="Best Fit")
        plt.text((concs[step]), 0.2, r'K$\mathbf{_d}$ = %.2f ± %.2f %s' % (popt, error, unit), fontweight='bold')
        plt.ylabel('R', fontweight='bold')
        plt.xlabel(r'[%s]$\mathbf{_0}$ (%s)' % (proteinName,unit), fontweight='bold')
        plt.xscale("log")
        plt.legend()
        plt.savefig("%s/bindingisotherm.png" % subdirect)           # save binding isotherm graph
        graphs.append(plt.figure())
        plt.close()

        # Statistics
        residuals = Rvalue - LevenMarqu(concs, *popt)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((Rvalue - np.mean(Rvalue))**2)
        r_squared = 1 - (ss_res/ss_tot)

        chiSquared = sum((((Rvalue - LevenMarqu(concs, *popt))**2) / LevenMarqu(concs, *popt)))


        ## Part 7 - Returning summary data and graphs
        # Summary dataframe of average signal per concentration with standard deviation, relative standard deviation, R value and standard deviation
        df = pd.DataFrame (forDF).transpose()
        df.columns = DFnames

        # Create new output sheet in input Excel file with summary data
        writer = pd.ExcelWriter(fileName, engine = 'openpyxl')
        writer.book = inputBook
        df.to_excel(writer, sheet_name = "Outputs", float_format='%.4f', index=False, startcol=3, engine = 'openpyxl')        # does not overwrite if a sheet named Outputs already exists
        writer.save()
        writer.close()

        outputSheet = inputBook.worksheets[len(inputBook.sheetnames)-1]

        # Duplicate input sheet information onto output sheet for reproducibility
        for r in range(1, 18):
                for c in range (1, 3):
                        outputSheet.cell(row=r, column=c).value = idealSheet.cell(row=r, column=c).value

        
        # Include Kd, R² and χ² values on output sheet
        outputSheet["K1"] = "Kd: %.4f ± %.4f %s" % (popt,error,unit)
        outputSheet["K2"] = "R²: %.4f" % (r_squared)
        outputSheet["K3"] = "χ²: %.4f" % (chiSquared)

        inputBook.save(fileName)

        # Testing script execution time
        end = time.time()
        print("Script run time: %.2f seconds" %(end-start))

        # Returning all graphs (separagrams and binding isotherm)
        #plt.show()
        plt.close('all')

        return subdirect




def report (workingFile, graphFolder):
    # Read in data tables from working file
    userInputs = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="A:B", engine='openpyxl')
    proteinName = str(userInputs.iloc[7,1])
    userLength = userInputs.shape[0]
    userInputs = userInputs.values.tolist()

    summaryTable = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="D:I", engine='openpyxl')
    summaryTable = summaryTable.dropna(how='any')
    sumLength = summaryTable.shape[0]
    summaryTable = summaryTable.values.tolist()

    kdTable = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="K", engine='openpyxl')
    kdTable = kdTable.dropna(how='any')
    kdTable = kdTable.values.tolist()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Times', size=10)

    pdf.cell(pdf.epw/2, pdf.font_size*2, '%s_%s' % (proteinName,date.today()))

    current = pdf.get_y() +5
    pdf.set_y(current)
    pdf.image("%s/bindingisotherm.png" % (graphFolder), w = pdf.epw/2)
    pdf.set_y(current)
    pdf.image("%s/allconcentration.png" % (graphFolder), w = pdf.epw/2, x = pdf.epw/2)

    # Add summary table to pdf
    current = pdf.get_y() + 5
    pdf.set_y(current)
    for row in summaryTable:
        pdf.set_x(2.5*(pdf.epw/5))
        
        for col in row:
            pdf.multi_cell(pdf.epw/12 , pdf.font_size*2, str(col), border=1,
                    new_x="RIGHT", new_y="TOP", align="L", max_line_height=pdf.font_size)
        pdf.ln(pdf.font_size*2)

    # Add Kd and statistics to pdf
    pdf.set_y(current + pdf.font_size*2*sumLength + 5)
    for row in kdTable:
        pdf.set_x(2.5*(pdf.epw/5))
        
        for col in row:
            if col.find('χ') != -1:
                col = col.replace("χ", "Chi")
            pdf.multi_cell(pdf.epw/4 , pdf.font_size*2, str(col), border=1,
                    new_x="RIGHT", new_y="TOP", align="L", max_line_height=pdf.font_size)
        pdf.ln(pdf.font_size*2)

    # Add user input table to pdf
    pdf.set_y(current)

    for row in userInputs:
        for col in row:
            pdf.multi_cell(pdf.epw/5, pdf.font_size*2, str(col), border=1,
                    new_x="RIGHT", new_y="TOP", align="L", max_line_height=pdf.font_size)
        pdf.ln(pdf.font_size*2)

    # Add separagram images to pdf
    images = natsorted(glob.glob('%s/*.png' % graphFolder))
    pdf.add_page()
    pos = 0.2
    graphY = pdf.get_y()
    count = 0

    for img in images:
        if img.endswith('M.png'):
            count +=1
            if count/16 == 1:
                pdf.add_page()
                count = 0
                
            if pos < 3:
                pdf.set_y(graphY)
                pdf.image(img, w = pdf.epw/3, x = pos*pdf.epw/3)
                pdf.ln(0)
                pos +=1
                
            elif pos == 3.2:
                pos = 0.2
                graphY = pdf.get_y()+5
                pdf.set_y(graphY)
                pdf.image(img, w = pdf.epw/3, x = pos*pdf.epw/3)
                pos = 1.2

    currentFolder = os.path.dirname(workingFile)
    pdfNAME = '%s/%s_%s' % (currentFolder, proteinName, date.today())
    pdf.output(pdfNAME)
    webbrowser.open_new('file://%s' % pdfNAME)







######### EVENT LOOP ##############
while True:
    event, values = window.read() 
    if event == sg.WIN_CLOSED:
        break

# Radiobuttons displaying relevant secondary fields
    # If compensation needed request concentration for normalization
    if event == 'compYes':
        window['normConc'].update(visible=True)
        window['normConc_val'].update(visible=True)
        
    if event == 'compNo':
        window['normConc'].update(visible=False)
        window['normConc_val'].update(visible=False)

    # If manual peak determination, enter peak times
    if event == 'peakM':
        window['manPeaks'].update(visible=True)
        window['manPeaks_val'].update(visible=True)
        window['progPeak'].update(visible=False)
        window['progPeak_val'].update(visible=False)
        
    # If programmatic peak determination, enter concentration to use
    if event == 'peakP':
        window['manPeaks'].update(visible=False)
        window['manPeaks_val'].update(visible=False)
        window['progPeak'].update(visible=True)
        window['progPeak_val'].update(visible=True)

# Validate user input file path
    if event == 'validate':
        window['out'].update(visible=False)
        window['report'].update(visible=False)

        # Confirm input is a string
        check = valStr(values['filePath_val'])
        if check[0] == True:
            filePath = check[1]
        elif check[0] == False:
            sg.popup(check[1])

        # Verify if input path exists
        if not os.path.exists(filePath):
            window['in'].update(visible=False)
            window['calculate'].update(disabled=True)

            window['invalid'].update(visible=True)
            window['work'].update(visible=False)
            window['rawData'].update(visible=False)
            
        # Verify if input path is Excel workbook or directory
        elif os.path.exists(filePath):
            if os.path.isfile(filePath) and filePath.endswith(".xlsx"):
                window['in'].update(visible=False)
                window['calculate'].update(disabled=False)

                window['invalid'].update(visible=False)
                window['work'].update(visible=True)
                window['rawData'].update(visible=False)
                pathType = "workingFile"
                
            elif os.path.isdir(filePath):
                window['in'].update(visible=True)
                window['calculate'].update(disabled=False)

                window['invalid'].update(visible=False)
                window['work'].update(visible=False)
                window['rawData'].update(visible=True)
                pathType = "directory"
   
# Validate user input and calculate Kd     
    if event == 'calculate':

        window['out'].update(visible=False)

        # If working file entered, validate contents
        if pathType == "workingFile":
                workingFile = str(values['filePath_val'])
                #### Validate workbook & sheets & values
                fileResults = validateExcel(workingFile)

                if fileResults[0]==False:
                    errorMessage = genFileErrorMessage(fileResults[1])
                    sg.popup(errorMessage)

                elif fileResults[0]:
                    sg.popup('Calculating Kd')
                
                data = pd.read_excel(workingFile, engine='openpyxl')
                inputBook = load_workbook(workingFile, data_only=True)         
                idealSheet = inputBook["Inputs"]
                compYN = str(idealSheet.cell(13,2).value)
                
        # If file path is directory validate required user inputs 
        elif pathType == "directory":
            filePath = str(values['filePath_val'])
            reqResult = confirmRequired(values)

            if reqResult[0]==False or reqResult[2]==False:
                errorMessage = genErrorMessage(reqResult[1], reqResult[3])
                sg.popup(errorMessage)

            elif reqResult[0] and reqResult[2]:
                sg.popup('Calculating Kd')

                # Read in user inputs for fluidics parameters
                propFlow = values['propFlow_val']
                injectFlow = values['injectFlow_val']
                injectTime = float(values['injectTime_val'])
                sepLength = values['sepLength_val']
                sepDiam = values['sepDiam_val']
                injectLength = values['injectLength_val']
                injectDiam = values['injectDiam_val']

                # Read in user inputs for concentrations
                proteinName = str(values['protName_val'])
                ligandName = values['ligName_val']
                ligandConc = float(values['ligConc_val'])

                # Read in user inputs for data analysis parameters
                if window['dataF'].get() == True:
                    dataType = "F"
                elif window['dataMS'].get() == True:
                    dataType = "MS"
                    
                if window['compYes'].get()== True:
                    compYN = "Y"
                    normalConc = float(values['normConc_val'])
                elif window['compNo'].get()== True:
                    compYN = "N"
                    normalConc = None

                windowWidth = values['window_val']

                if values['peakM'] == True:
                    peakDet = "M"
                    manualPeaks = values['manPeaks_val']
                    peakConc=None
                elif values['peakP'] == True:
                    peakDet = "P"
                    manualPeaks = None
                    peakConc = float(values['progPeak_val'])

            
                # Check for existing Excel files with same name
                suggName = "%s/%s_%s.xlsx" % (filePath, proteinName, date.today())
                exists = checkForOverwriting(suggName)

                if exists[0] == True:
                    if sg.popup_yes_no('A file with the file path %s already exists in this directory. \n \n Would you like to overwrite this file?' % os.path.basename(suggName)) =='NO':
                        workingFilePath = exists[1]
                        
                    else:
                        os.remove(suggName)
                        workingFilePath = (suggName)
                        
                    
                elif exists[0] == False:
                    workingFilePath = suggName
        
                # Validate contents of raw data directory
                checkDirect = validateDirectoryContents(filePath, compYN, normalConc, peakDet, peakConc)

                if checkDirect[0] == False:
                    errorMessage = genErrorMessageDirect(checkDirect[1])
                    sg.popup(errorMessage)
                    window.close()
                    break
                    
                # Prepare working file
                workingFile = workingfileprep(filePath, workingFilePath, propFlow, injectFlow, injectTime, sepLength, sepDiam,
                                              injectLength, injectDiam, proteinName, ligandName, ligandConc, dataType,
                                              compYN, normalConc, windowWidth, peakDet, manualPeaks, peakConc)


        # Unmask signals with compensation procedure if indicated
        if compYN == "Y" :
            compensate(workingFile)
                

        # Analyze signals, update Excel working file and generate subfolder with separagrams and binding isotherm
        graphPath = dataanalysis(workingFile)
        images = natsorted(glob.glob('%s/*.png' % graphPath))
        load_image(images[0],window)
            
        # Read in summary and Kd information from working file to display in GUI output
        df = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="D:I", engine='openpyxl')
        df = df.dropna(how='any')
        headers = df.iloc[0].values.tolist()
        data = df.iloc[1:].values.tolist()
        window['summary'].update(values=data, num_rows=min(10,len(data)))

        df = pd.read_excel(workingFile, sheet_name=-1, header=None, usecols="K", engine='openpyxl')
        df = df.dropna(how='any')
        data = df.values.tolist()
        window['Kd'].update(values=data, num_rows=3)
            
        # Show output column and report
        window['out'].update(visible=True)
        window['report'].update(visible=True)

        
    # Buttons for image viewer in GUI
    if event == 'fwd':
        if location == len(images)-1:
            location=0
        else:
            location +=1
        load_image(images[location], window)

    if event == 'back':
        if location == 0:
            location=len(images)-1
        else:
             location -=1
        load_image(images[location], window)

    # Functionality for report button to generate, save and open PDF report
    if event == 'report':
        report(workingFile, graphPath) 

window.close()
