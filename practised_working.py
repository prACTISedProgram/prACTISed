#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# practised_working.py generates an Excel working file from ACTIS
# experimental data for use with practised.py

# Copyright (C) 2022  Jessica Latimer

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

def workingfileprep(inputPath, workingFilePath, propFlow, injectFlow, injectTime, sepLength, sepDiam,
                    injectLength, injectDiam, proteinName, ligandName, ligandConc,
                    dataType, compYN, normalConc, windowWidth, peakDet, manualPeaks, peakConc):

    
    ## Part 1 - Importing the required libraries and sub-libraries required below
    import argparse                                   
    import pathlib
    import sys
    import pandas as pd
    import matplotlib.pyplot as plt
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils import FORMULAE
    from openpyxl.utils import get_column_letter, column_index_from_string
    import math
    from scipy.optimize import curve_fit

    import os
    from csv import reader
    from natsort import natsorted
    from datetime import date

    from practised_knuteon import readTrace, variant_to_system_time, read_traces, extract_trace
    import practised_pwexplode
    

    d = {}
    timeRun = {}
    prefix=""

    def isfloat(num):
        try:
            float(num)
            return True
        except ValueError:
            return False     

    ### Reading in files from directory ###
    for file in os.listdir(inputPath):

        ## Read in simulated protein profile if compensation procedure was selected
        if compYN == "Y" and file.endswith((".txt")) and file.startswith('simulated'):

             # Extract the preamble information
             preamble = []
             with open("%s/%s" %(inputPath, file), "r", encoding="latin-1") as fileCheck:
                 csv_reader = reader(fileCheck, delimiter=" ")
         
                 for row in csv_reader:
                     if row[0].isalpha() == True or isfloat(row[0])==False:
                         preamble.append(row)
                           
                     elif row[0].isalpha() == False or isfloat(row[0])==True:
                         break

             # Read in simulated protein profile if compensation procedure was selected
             simulated = pd.read_csv("%s/%s" % (inputPath,file), sep="\s+", encoding="latin-1", skiprows=len(preamble), header=None, keep_default_na=True, na_values=str(0))
             simulated = simulated.fillna(0)
             simulated.columns = ['raw time', 'signal']

             # Normalize simulated protein signals
             maxSim = simulated['signal'].max()
             simulated['signal'] = simulated['signal'].div(maxSim)


        ## Read in raw data files 
        if file.endswith((".txt", ".asc", ".dat")) and not file.startswith('simulated') and not 'READ' in os.path.splitext(file)[0]:

            # Extract concentration prefix
            molar = file.find("M")
            prefix = file[molar-1]

            # Extract concentration and run number from file name (see naming conventions)
            conc = float(file.partition(prefix)[0])

            name = os.path.splitext(file)[0]
            if name[-1].isdigit()==True: 
                 if name[-2:].isdigit()==True: 
                    runNumber= int(name[-2:])
                 else:
                    runNumber= int(name[-1])
            else:
                runNumber=1


            # Extract trace from .dat files using modified knuteon code
            if file.endswith(".dat"):
                run = readTrace("%s/%s" % (inputPath, file))
                run.columns = ["raw time", "Experiment " + str(runNumber)]
                
                if conc in d:
                    d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run["Experiment " + str(runNumber)])

                    if runNumber < timeRun[conc]:
                            d[conc].loc[:,'raw time'] = run.iloc[:,0]
                            timeRun[conc] = runNumber

                elif conc not in d:
                    d[conc]=run
                    timeRun[conc] = runNumber
                
            # Extract the preamble information from .asc or .txt files
            elif file.endswith((".txt", ".asc")):
                preamble = []
                delim = "\t"
                with open("%s/%s" %(inputPath, file), "r", encoding="latin-1") as fileCheck:
                    csv_reader = reader(fileCheck, delimiter= delim)
                    for row in csv_reader:
                           
                        if row[0].isalpha() == True or isfloat(row[0])==False:
                            preamble.append(row)
                            
                        elif row[0].isalpha() == False or isfloat(row[0])==True:
                            break
                        
                if len(preamble) > 20:
                    preamble = []
                    delim = "\s+"
                    with open("%s/%s" %(inputPath, file), "r", encoding="latin-1") as fileCheck:
                        csv_reader = reader(fileCheck, delimiter=" ")
                        for row in csv_reader:
                               
                            if row[0].isalpha() == True or isfloat(row[0])==False:
                                preamble.append(row)
                                
                            elif row[0].isalpha() == False or isfloat(row[0])==True:
                                break
                    
                # If no multiplier extract time and signal
                if len(preamble) <= 1:
                    run = pd.read_csv("%s/%s" % (inputPath,file), sep= delim, encoding="latin-1", keep_default_na=True, na_values=str(0))
                    run = run.dropna(how="all")
                    run = run.fillna(0)
                    run = run.iloc[:,[0,1]]
                    run.columns = ["raw time", "Experiment " + str(runNumber)]
                    if '(s)' not in preamble[0]:
                        run.iloc[:,0] = run.iloc[:,0].mul(60)


                    # Create or add experiment to dataframe if it exists
                    if conc in d:
                        d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run["Experiment " + str(runNumber)])

                        if runNumber < timeRun[conc]:
                            d[conc].loc[:,'raw time'] = run.iloc[:,0]
                            timeRun[conc] = runNumber
                            
                    elif conc not in d:
                        d[conc]=run
                        timeRun[conc] = runNumber

                # If multiplier needed extract signals, signal multipler, and iterate over signals
                elif len(preamble) > 1:
                    signalMult_line = list(filter(lambda x: "Y Axis Multiplier:" in x[0], preamble))
                    signalMult = float(signalMult_line[0][1])

                    run = pd.read_csv("%s/%s" %(inputPath, file), sep="\t", encoding="latin-1", skiprows=len(preamble), header=None, keep_default_na=True, na_values=str(0))
                    run = run.dropna(how="all")
                    run = run.fillna(0)
                    run.columns = ["Experiment " + str(runNumber)]
                    run.loc[:"Experiment " + str(runNumber)] *= signalMult

                    # Create or add experiment to dataframe if it exists
                    if conc in d:
                        d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run)

                    elif conc not in d:
                        # Reconstruct raw time from sampling rate
                        hz_line = list(filter(lambda x: "Sampling Rate:" in x[0], preamble))
                        hz = float(hz_line[0][1])
                        second_Gap = 1/hz

                        rawTime = [0]

                        for x in range(0,len(run)-1):
                            rawTime.append(rawTime[x]+second_Gap)

                        run.insert(0, "raw time", rawTime)
                        d[conc]=run

                            
    for xConc in d:
        orderedExp = natsorted(list(d[xConc].columns))
        orderedExp.remove("raw time")
        orderedExp.insert(0,"raw time")
        d[xConc] = d[xConc].reindex(columns = orderedExp)

            
    orderedDict = natsorted(d.keys())

    if prefix == "u":
        prefix = "µ"

    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"

    writer = pd.ExcelWriter(workingFilePath, engine = 'openpyxl')
    writer.book = wb
        
    # Generate Inputs Sheet
    inputDictionary = {"Propogation flow rate":propFlow, "Injection flow rate":injectFlow, "Injection time (s)":injectTime,
                       "Separation capillary length":sepLength, "Separation capillary diameter":sepDiam,
                       "Injection loop length":injectLength, "Injection loop diameter":injectDiam, "Protein name":proteinName,
                       "Ligand name":ligandName, "Number of Concentrations":len(d.keys()),
                       "Initial Ligand concentration [L]0 (%sM)" % prefix: ligandConc, "Type of Data":dataType,
                       "Compensation procedure": compYN, "[P]0 reference for MS normalization (%sM)" % prefix: normalConc,
                       "Window width (%)": windowWidth, "Determination of peak": peakDet, "Manually determined peaks":manualPeaks,
                       "[P]0 to programmaticlly determine peak":peakConc
                       }

    for key in inputDictionary:
        row = list(inputDictionary.keys()).index(key)+1
        ws["A%s" % row] = key
        ws["B%s" % row] = inputDictionary[key]
                   
    for x in range(1,len(d)+1):
        ws["D"+str(x)] = "Protein Conc. #" + str(x)
        ws["E"+str(x)] = "%s %sM" % (orderedDict[x-1], prefix)


   # Add simulated protein profile if compensation required
    if compYN == "Y":
        simulated.to_excel(writer, sheet_name = "P_simulated", index=False)

    # Add sheet for each concentration dataframe
    for y in orderedDict:
        d[y].to_excel(writer, sheet_name = "%s %sM" % (y, prefix), index=False)
            
    writer.save()
    writer.close()
    wb.save(workingFilePath)

    return workingFilePath
