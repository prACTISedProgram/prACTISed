
# First attempt at Working File Preparation from .asc to Excel
# Makes a working file like idealinputs.xlsx, but called testExcel.xlsx

#20220716 JL    NOTE: natsort dependecy - variables to be passed from GUI, potentially add to GUI entries
#20220720 JL    NOTE: adding all experimental parameters

## Part 1 - Importing the required libraries and sub-libraries required below
import argparse                                   
import pathlib
import sys
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from scipy.optimize import curve_fit

import os
from csv import reader
from natsort import natsorted

# Testing script execution time
import time
start = time.time()


### NOTE: data from GUI required to populate Excel working file inputs sheet ###
    #numConc
    #window
    #injectTime
    #ligConc
    #windowConc

workingFileName = "testExcel"                       # Would normally be passed from GUI
directory = "/Users/jess/Documents/practised"       # Would normally be passed from GUI
#fileType = ".asc"                                  # Only supports .asc files - need to include .txt and .dat files

d = {}

### Reading in raw data files ###
for file in os.listdir(directory):

    if file.endswith(".asc"):

        if not pathlib.Path(file).is_file():
            print("Given file '%s' is not a file or does not exist." % file)
            exit(-1)
    
        preamble =[]

        # Extract the preamble information
        with open(file, "r", encoding="latin-1") as fileCheck:
           csv_reader = reader(fileCheck, delimiter="\t")
           for row in csv_reader:
               if row[0].isnumeric() == False:
                   preamble.append(row)
                   
               else:
                   exit
                   
        # Extract concentration and run number
        sample_line = list(filter(lambda x: "Sample ID" in x[0], preamble))
        conc = float(sample_line[0][1].partition("u")[0])

        if file[-5].isnumeric:                  # Assumption about file name format
            runNumber= int(file[-6:-4])         # Update prACTISed.py to handle when runs do not start at 1

        else:
            runNumber=1

        # Extract signals, signal multipler, and iterate over signals
        signalMult_line = list(filter(lambda x: "Y Axis Multiplier:" in x[0], preamble))
        signalMult = float(signalMult_line[0][1])

        run = pd.read_csv(file, sep=",", encoding="latin-1", skiprows=len(preamble), header=None, keep_default_na=True, na_values=str(0))
        run = run.dropna(how="any")
        run.columns = ["Experiment" + str(runNumber)]
        run.loc[:"Experiment" + str(runNumber)] *= signalMult

        # Add Experiment to dataframe if it exists, or create data frame and add raw time
        if conc in d:
            d[conc].insert(len(d[conc].columns), "Experiment " + str(runNumber), run)
        
        elif conc not in d:
            
            hz_line = list(filter(lambda x: "Sampling Rate:" in x[0], preamble))
            hz = float(hz_line[0][1])
            second_Gap = 1/hz

            rawTime = [0]

            for x in range(0,len(run)-1):
                rawTime.append(rawTime[x]+second_Gap)

            run.insert(0, "raw time", rawTime)
            d[conc]=run

for xConc in d:
    orderedExp = natsorted(list(d[xConc].columns))
    orderedExp.remove("raw time")
    orderedExp.insert(0,"raw time")
    d[xConc] = d[xConc].reindex(columns = orderedExp)
    
orderedDict = natsorted(d.keys())

# Create Excel Working File
wb = Workbook()
ws = wb.active

# Generate Inputs Sheet
ws.title = "Inputs"
ws["A1"] = "Propogation flow rate (µM/min)"
ws["B1"] = "TBD from GUI"
ws["A2"] = "Injection flow rate (µM/min)"
ws["B2"] = "TBD from GUI"
ws["A3"] = "Injection time"
ws["B3"] = "TBD from GUI"
ws["A4"] = "Separation capillary length (cm)"
ws["B4"] = "TBD from GUI"
ws["A5"] = "Injection loop length (cm)"
ws["B5"] = "TBD from GUI"
ws["A6"] = "Protein name"
ws["B6"] = "TBD from GUI"
ws["A7"] = "Ligand name"
ws["B7"] = "TBD from GUI"
ws["A8"] = "Number of Concentrations"
ws["B8"] = "TBD from GUI"
ws["A9"] = "Initial Ligand concentration [L]0 (µM)"
ws["B9"] = "TBD from GUI"
ws["A10"] = "Type of data (MS) or (F)"
ws["B10"] = "TBD from GUI"
ws["A11"] = "[P]0 reference for MS normalization (µM)"
ws["B11"] = "TBD from GUI"
ws["A12"] = "Window width (%)"
ws["B12"] = "TBD from GUI"
ws["A13"] = "Determination of peak (M) or (P)"
ws["B13"] = "TBD from GUI"
ws["A14"] = "Manually determined peaks"
ws["B14"] = "TBD from GUI" 
ws["A15"] = "[P]0 used to programmaticlly determine peak"
ws["B15"] = "TBD from GUI"               


for x in range(1,len(d)+1):
    ws["D"+str(x)] = "Protein Conc. #" + str(x)
    ws["F"+str(x)] = "Runs"

    ws["E"+str(x)] = orderedDict[x-1]
    ws["G"+str(x)] = len(d[orderedDict[x-1]].columns)-1

# Write each concentration to a new sheet with all runs and raw time
writer = pd.ExcelWriter("testExcel.xlsx", engine = 'openpyxl')          ## User could name Excel file in GUI
writer.book = wb

for y in orderedDict:
    d[y].to_excel(writer, sheet_name = str(float(y)) + " µM", index=False)
    
writer.save()
writer.close()
wb.save("%s.xlsx" % workingFileName)

# Testing script execution time
end = time.time()
print("Script run time: %.2f seconds" %(end-start))
